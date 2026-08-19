"""
Sourcing Pipeline — Local Web UI (Automation-first redesign)
============================================================
Non-technical UI for the scraping + Keepa + sourcing pipeline.

Client workflow:
  1. Pick supplier (Zoro or iHerb)
  2. Pick how many products to analyze
  3. Pick approval criteria (preset)
  4. Click Start
  5. Download the resulting xlsx

Advanced filter controls are tucked into a "Adjust individual filters"
disclosure — hidden by default so the client doesn't see them unless
they want to.

Usage:
    pip install flask openpyxl
    python webapp.py
"""

from __future__ import annotations

import json
import logging
import os
import re
import secrets
import subprocess
import sys
import threading
import time
import uuid
from datetime import timedelta
from functools import wraps
from pathlib import Path

from flask import (
    Flask, abort, jsonify, redirect, render_template_string,
    request, send_file, session, url_for,
)

from sourcing import SourcingConfig

# ---------------------------------------------------------------- #
# Setup                                                            #
# ---------------------------------------------------------------- #

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("webapp")

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False

BASE_DIR = Path(__file__).resolve().parent
WORK_DIR = BASE_DIR / ".webapp_runs"
WORK_DIR.mkdir(exist_ok=True)

JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


# ---------------------------------------------------------------- #
# Auth: signed-cookie sessions (no database, ~microseconds/req)    #
# ---------------------------------------------------------------- #

# SECRET_KEY signs the session cookie. Must be stable across restarts, otherwise
# every user gets logged out on redeploy. setup_vps.sh generates one and writes
# it to .env. If missing (dev), fall back to an ephemeral one and warn.
_secret = os.environ.get("SECRET_KEY", "").strip()
if not _secret:
    log.warning("SECRET_KEY not set — using ephemeral key. Sessions won't "
                "survive restart. Set SECRET_KEY in .env for production.")
    _secret = secrets.token_hex(32)
app.config["SECRET_KEY"] = _secret
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# Set SESSION_COOKIE_SECURE=1 in .env once you're behind HTTPS.
if os.environ.get("SESSION_COOKIE_SECURE", "").lower() in ("1", "true", "yes"):
    app.config["SESSION_COOKIE_SECURE"] = True

APP_USERNAME = os.environ.get("APP_USERNAME", "").strip()
APP_PASSWORD = os.environ.get("APP_PASSWORD", "").strip()
AUTH_ENABLED = bool(APP_USERNAME and APP_PASSWORD)
if not AUTH_ENABLED:
    log.warning("APP_USERNAME/APP_PASSWORD not set — auth is DISABLED (dev "
                "mode). Set both in .env for production.")


def require_login(fn):
    """Redirect HTML routes to /login, return 401 JSON for /api/*."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not AUTH_ENABLED or session.get("user") == APP_USERNAME:
            return fn(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "unauthorized"}), 401
        return redirect(url_for("login", next=request.path))
    return wrapper


# ---------------------------------------------------------------- #
# Presets — three clear approval-strictness levels                 #
# ---------------------------------------------------------------- #

PRESETS = {
    "balanced": {
        "label": "Balanced",
        "sub": "Your standard rules",
        "desc": "At least 4 FBM sellers, 3.5+ rating, positive profit with 10% margin.",
        "values": {
            "min_profit": 2, "min_margin": 0.10, "min_roi": 0.15,
            "min_fbm_sellers": 4, "min_rating": 3.5, "min_reviews": 0,
            "require_rating": False, "require_reviews": False,
            "reject_if_amazon_on_listing": False, "max_bsr": 0,
            "min_historical_sellers": 0,
        },
    },
    "loose": {
        "label": "Discovery",
        "sub": "See more, filter less",
        "desc": "Almost no filtering. Good for exploring what's out there before narrowing.",
        "values": {
            "min_profit": 0.01, "min_margin": 0, "min_roi": 0,
            "min_fbm_sellers": 0, "min_rating": 0, "min_reviews": 0,
            "require_rating": False, "require_reviews": False,
            "reject_if_amazon_on_listing": False, "max_bsr": 0,
            "min_historical_sellers": 0,
        },
    },
    "strict": {
        "label": "Amazon-safe",
        "sub": "Only clear winners",
        "desc": "Requires 50+ reviews, 4.0+ rating, healthy BSR, and 15% margin. Fewer, more confident.",
        "values": {
            "min_profit": 5, "min_margin": 0.15, "min_roi": 0.25,
            "min_fbm_sellers": 4, "min_rating": 4.0, "min_reviews": 50,
            "require_rating": True, "require_reviews": True,
            "reject_if_amazon_on_listing": False, "max_bsr": 100000,
            "min_historical_sellers": 5,
        },
    },
}

RETAILERS = [
    {
        "value": "https://www.zoro.com",
        "label": "Zoro",
        "sub": "US supplier | prices in USD",
    },
    {
        "value": "https://www.iherb.com",
        "label": "iHerb",
        "sub": "supplements | US pricing forced",
    },
]


# ---------------------------------------------------------------- #
# Job execution                                                    #
# ---------------------------------------------------------------- #

def _make_job(mode: str) -> str:
    job_id = uuid.uuid4().hex[:12]
    with JOBS_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "mode": mode,
            "status": "queued",
            "stage": "starting",
            "log": [],
            "error": None,
            "output_file": None,
            "summary": None,
            "created_at": time.time(),
            "keepa_exhausted": False,   # sticky flag; set from log scan
            "keepa_tokens_left": None,  # last known token balance
            # Live progress fields (populated by _append_log regex scan)
            "current": None,             # e.g. 15
            "total": None,               # e.g. 25
            "current_activity": None,    # human-readable line, e.g. "Keepa UPC 12345"
            "urls_discovered": None,     # from "Sitemap yielded N URL(s)"
            "urls_after_filter": None,   # from "N remain for scraping"
            "eta_seconds": None,         # simple linear extrapolation
            # ETA baseline — snapshotted on first N/M seen in current stage
            "_eta_stage": None,
            "_eta_start_time": None,
            "_eta_start_count": None,
        }
    return job_id


# --- Regex anchors for log-line progress extraction ---
# These patterns match lines emitted by scraper.py and sourcing.py. Keep in
# sync with those modules — if you rename a log message there, update here.
# NOTE: scraper.py's logger prepends "HH:MM:SS [INFO] " to every line, so
# _RE_STEP uses `search` (not `match`) with a whitespace anchor before `[N/M]`
# to avoid false positives on nested "[INFO]"-style prefixes.
_RE_STEP = re.compile(r"(?:^|\s)\[(\d+)/(\d+)\]\s+(.+?)(?:\s*$)")
_RE_URLS_YIELDED = re.compile(r"Sitemap yielded (\d+) URL", re.IGNORECASE)
_RE_URLS_KEPT = re.compile(r"(\d+) remain(?:ing)?(?:\s+for scraping)?",
                           re.IGNORECASE)
_RE_TOKENS = re.compile(r"Keepa tokens remaining:\s*(\d+)", re.IGNORECASE)
_RE_FETCHING = re.compile(r"Fetching\s+(https?://\S+)")


def _summarize_step(rest: str) -> str:
    """
    Turn a raw '[N/M] ...' line body into a compact human-readable
    'current activity' string for the progress card. Kept short — the UI
    shows this inline next to the progress bar.
    """
    if "Keepa UPC" in rest:
        upc = rest.split("Keepa UPC", 1)[1].strip()[:20]
        return f"Amazon lookup for UPC {upc}"
    if "Keepa title search" in rest:
        after = rest.split("Keepa title search", 1)[1].strip()
        return f"Amazon title search: {after[:50]}"
    if "Scraping product" in rest:
        return "Scraping supplier product page"
    return rest[:80]


def _update_eta(job: dict, current: int, total: int) -> None:
    """
    Simple linear-rate ETA. Snapshots progress on entry to each stage so
    a slow first item doesn't skew the estimate for the whole run.
    """
    if total <= 0 or current <= 0:
        return
    stage = job.get("stage")
    now = time.time()
    # Reset baseline whenever stage changes so ETA is per-stage.
    if job.get("_eta_stage") != stage:
        job["_eta_stage"] = stage
        job["_eta_start_time"] = now
        job["_eta_start_count"] = current - 1  # count as if we just started
    elapsed = now - (job["_eta_start_time"] or now)
    done_since_baseline = current - (job["_eta_start_count"] or 0)
    if elapsed >= 3 and done_since_baseline > 0:
        rate = done_since_baseline / elapsed  # items/sec
        remaining = max(0, total - current)
        if rate > 0:
            job["eta_seconds"] = int(remaining / rate)


def _append_log(job_id: str, line: str) -> None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is None:
            return
        job["log"].append(line)
        if len(job["log"]) > 500:
            job["log"] = job["log"][-500:]

        low = line.lower()
        if "sitemap" in low or "discovered" in low or "collected" in low:
            job["stage"] = "discovering"
        elif "scraping product" in low or "fetching " in low:
            job["stage"] = "scraping"
        elif "keepa" in low and ("lookup" in low or "upc" in low or "title search" in low):
            job["stage"] = "matching"
        elif "sourcing done" in low or "wrote " in low and "sourcing rows" in low:
            job["stage"] = "finalizing"

        # Surface Keepa quota-exhaustion state to the UI so users see a
        # dedicated warning banner instead of just a confusing "some rows
        # rejected" message.
        if "keepa quota exhausted" in low or "keepa tokens exhausted" in low:
            job["keepa_exhausted"] = True

        # --- Structured progress extraction (per-item counters, ETA) ---

        # [N/M] step progress — anchor for scraping AND matching stages.
        # `search` is intentional: log lines are prefixed with a timestamp +
        # "[INFO]" by the scraper subprocess before hitting stdout.
        m = _RE_STEP.search(line)
        if m:
            try:
                current, total = int(m.group(1)), int(m.group(2))
                job["current"] = current
                job["total"] = total
                job["current_activity"] = _summarize_step(m.group(3))
                _update_eta(job, current, total)
            except (ValueError, IndexError):
                pass

        # Discovery counters — tell the user how many URLs were found
        m = _RE_URLS_YIELDED.search(line)
        if m:
            try:
                job["urls_discovered"] = int(m.group(1))
                job["current_activity"] = f"Discovered {m.group(1)} product URLs"
            except (ValueError, IndexError):
                pass
        m = _RE_URLS_KEPT.search(line)
        if m and "URL pre-filter" in line:
            try:
                job["urls_after_filter"] = int(m.group(1))
            except (ValueError, IndexError):
                pass

        # Live token balance (heartbeat lines from sourcing.py + final line)
        m = _RE_TOKENS.search(line)
        if m:
            try:
                job["keepa_tokens_left"] = int(m.group(1))
            except (ValueError, IndexError):
                pass

        # Also useful: individual Fetching lines when [N/M] wasn't present
        # (legacy scraper.scrape_one direct callers)
        if job.get("current_activity") is None:
            fm = _RE_FETCHING.search(line)
            if fm:
                url = fm.group(1)
                slug = url.rstrip("/").split("/")[-2 if url.endswith("/") else -1]
                job["current_activity"] = f"Fetching {slug[:60]}"


def _mark_status(job_id: str, status: str, **extra) -> None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is not None:
            job["status"] = status
            for k, v in extra.items():
                job[k] = v


def _run_scraper_subprocess(job_id: str, args: list[str]) -> None:
    log.info("Job %s: %s", job_id, " ".join(args))
    _mark_status(job_id, "running")
    try:
        proc = subprocess.Popen(
            [sys.executable, "-u", "scraper.py", *args],
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        for line in proc.stdout:  # type: ignore[union-attr]
            _append_log(job_id, line.rstrip())
        proc.wait()
        if proc.returncode != 0:
            # Grab the last few log lines so the UI can show a useful error
            # instead of just "check the log"
            with JOBS_LOCK:
                tail = "\n".join(JOBS[job_id]["log"][-15:]) if JOBS.get(job_id) else ""
            if proc.returncode == 2:
                err = (
                    "The scraper rejected an argument (exit code 2). "
                    "This usually means scraper.py is an older version "
                    "missing a required flag. Verify with: "
                    "python scraper.py --help | Select-String \"config\"\n\n"
                    "Last log lines:\n" + tail
                )
            else:
                err = (f"Pipeline exited with code {proc.returncode}.\n\n"
                       f"Last log lines:\n{tail}")
            _mark_status(job_id, "error", error=err)
            return
    except FileNotFoundError as e:
        _mark_status(job_id, "error", error=f"scraper.py not found: {e}")
        return
    except Exception as e:  # noqa: BLE001
        _mark_status(job_id, "error", error=str(e))
        return

    _mark_status(job_id, "summarizing")


def _summarize_xlsx(job_id: str, xlsx_path: Path) -> None:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[2]]

        def col(name: str) -> int | None:
            try:
                return headers.index(name)
            except ValueError:
                return None

        c_status = col("Status")
        c_reasons = col("Reject Reasons")
        c_title = col("Supplier Title")
        c_profit = col("Profit $")
        c_roi = col("ROI %")
        c_asin = col("Amazon ASIN")
        c_supplier_cost = col("Supplier Cost (USD)")
        c_amazon_price = col("Amazon Sell Price")
        c_fbm = col("FBM Sellers (live)")

        by_status: dict[str, int] = {}
        approved_rows: list[dict] = []
        reason_tally: dict[str, int] = {}

        for row in ws.iter_rows(min_row=3, values_only=True):
            if c_status is None:
                continue
            status = row[c_status] or "Unknown"
            by_status[status] = by_status.get(status, 0) + 1

            if status == "APPROVED":
                approved_rows.append({
                    "title": (row[c_title] or "") if c_title is not None else "",
                    "profit": row[c_profit] if c_profit is not None else None,
                    "roi": row[c_roi] if c_roi is not None else None,
                    "asin": row[c_asin] if c_asin is not None else None,
                    "supplier_cost": row[c_supplier_cost] if c_supplier_cost is not None else None,
                    "amazon_price": row[c_amazon_price] if c_amazon_price is not None else None,
                    "fbm": row[c_fbm] if c_fbm is not None else None,
                })

            if c_reasons is not None and status == "Rejected":
                for reason in (row[c_reasons] or "").split(";"):
                    reason = reason.strip()
                    if not reason:
                        continue
                    if "FBM sellers" in reason:
                        key = "Not enough FBM sellers"
                    elif "profit" in reason.lower():
                        key = "Profit too low"
                    elif "Amazon is on" in reason:
                        key = "Amazon competes on the listing"
                    elif "rating" in reason.lower():
                        key = "Rating too low"
                    elif "reviews" in reason.lower():
                        key = "Not enough reviews"
                    elif "BSR" in reason:
                        key = "Selling too slowly (BSR)"
                    elif "no Amazon match" in reason:
                        key = "Not on Amazon"
                    elif "out of stock" in reason.lower():
                        key = "Out of stock at supplier"
                    elif "margin" in reason.lower():
                        key = "Margin too thin"
                    elif "ROI" in reason:
                        key = "ROI too low"
                    else:
                        key = reason[:60]
                    reason_tally[key] = reason_tally.get(key, 0) + 1

        approved_rows.sort(
            key=lambda r: r["profit"] if isinstance(r["profit"], (int, float)) else -999,
            reverse=True,
        )

        summary = {
            "total": sum(by_status.values()),
            "by_status": by_status,
            "top_approved": approved_rows[:10],
            "reject_reasons": sorted(reason_tally.items(), key=lambda kv: -kv[1])[:8],
        }
        _mark_status(job_id, "done",
                     output_file=str(xlsx_path.resolve()),
                     summary=summary)
    except Exception as e:  # noqa: BLE001
        log.exception("Summary error for job %s", job_id)
        _mark_status(job_id, "error", error=f"Summary failed: {e}")


def _job_scrape_and_analyze(job_id: str, params: dict) -> None:
    try:
        cfg_path = WORK_DIR / f"{job_id}_config.json"
        cfg_path.write_text(json.dumps(params["config"], indent=2), encoding="utf-8")

        out_path = WORK_DIR / f"{job_id}_result.xlsx"

        args = [
            "--sitemap", params["sitemap_url"],
            "--limit", str(params["limit"]),
            "--sourcing",
            "--config", str(cfg_path),
            "--out", str(out_path),
            "--no-robots",
        ]
        if params.get("random"):
            args.append("--random")
        if params.get("min_supplier_price", 0) > 0:
            args.extend(["--min-supplier-price", str(params["min_supplier_price"])])
        if params.get("brands"):
            args.extend(["--brands", params["brands"]])

        _run_scraper_subprocess(job_id, args)
        if JOBS[job_id]["status"] == "error":
            return
        _summarize_xlsx(job_id, out_path)
    except Exception as e:  # noqa: BLE001
        log.exception("Job %s failed", job_id)
        _mark_status(job_id, "error", error=str(e))


# ---------------------------------------------------------------- #
# Routes                                                           #
# ---------------------------------------------------------------- #

@app.route("/login", methods=["GET", "POST"])
def login():
    if not AUTH_ENABLED:
        return redirect(url_for("index"))
    if session.get("user") == APP_USERNAME:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = request.form.get("password") or ""
        remember = request.form.get("remember") == "on"
        # secrets.compare_digest is constant-time — resists timing attacks.
        if (secrets.compare_digest(u, APP_USERNAME)
                and secrets.compare_digest(p, APP_PASSWORD)):
            session.clear()
            session["user"] = APP_USERNAME
            session.permanent = bool(remember)
            next_url = request.args.get("next", "/")
            if not next_url.startswith("/") or next_url.startswith("//"):
                next_url = "/"
            return redirect(next_url)
        # Small delay makes brute-force annoying without hurting real UX.
        time.sleep(0.6)
        error = "That username or password isn't right. Try again."

    return render_template_string(LOGIN_HTML, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@require_login
def index():
    return render_template_string(
        INDEX_HTML,
        presets=PRESETS,
        retailers=RETAILERS,
        username=session.get("user") or "",
        auth_enabled=AUTH_ENABLED,
    )


@app.route("/api/run", methods=["POST"])
@require_login
def api_run():
    data = request.get_json(force=True)
    required = ["sitemap_url", "limit", "config"]
    missing = [k for k in required if k not in data]
    if missing:
        return jsonify({"error": f"Missing: {missing}"}), 400
    job_id = _make_job("scrape")
    threading.Thread(
        target=_job_scrape_and_analyze,
        args=(job_id, data), daemon=True,
    ).start()
    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
@require_login
def api_status(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is None:
            return jsonify({"error": "unknown job"}), 404
        return jsonify({
            "status": job["status"],
            "stage": job["stage"],
            "log_tail": job["log"][-25:],
            "log_count": len(job["log"]),
            "error": job["error"],
            "summary": job["summary"],
            "keepa_exhausted": job.get("keepa_exhausted", False),
            "keepa_tokens_left": job.get("keepa_tokens_left"),
            "current": job.get("current"),
            "total": job.get("total"),
            "current_activity": job.get("current_activity"),
            "urls_discovered": job.get("urls_discovered"),
            "urls_after_filter": job.get("urls_after_filter"),
            "eta_seconds": job.get("eta_seconds"),
        })


@app.route("/api/download/<job_id>")
@require_login
def api_download(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if job is None or not job.get("output_file"):
        abort(404)
    return send_file(
        job["output_file"],
        as_attachment=True,
        download_name=f"sourcing_{job_id}.xlsx",
    )


# ---------------------------------------------------------------- #
# HTML template                                                    #
# ---------------------------------------------------------------- #

LOGIN_HTML = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Sign in - Sourcing</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --paper:        #070B14;
      --card:         rgba(18, 26, 42, 0.88);
      --border:       rgba(255, 255, 255, 0.08);
      --border-strong:rgba(255, 255, 255, 0.16);
      --ink:          #F2F5FA;
      --graphite:     #B7C0D1;
      --muted:        #7D8AA3;
      --brand:        #E8C547;
      --brand-hover:  #F3D56A;
      --brand-soft:   rgba(232, 197, 71, 0.16);
      --rejected:     #F07167;
      --rejected-soft:rgba(240, 113, 103, 0.14);
      --sans:  "Plus Jakarta Sans", Inter, system-ui, sans-serif;
      --display: "Outfit", "Plus Jakarta Sans", system-ui, sans-serif;
    }

    * { box-sizing: border-box; }
    html, body { margin: 0; padding: 0; height: 100%; }
    body {
      font-family: var(--sans);
      color: var(--ink);
      line-height: 1.5;
      color-scheme: dark;
      -webkit-font-smoothing: antialiased;
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 24px;
      background:
        radial-gradient(900px 520px at 10% -20%, rgba(99, 102, 241, 0.28), transparent 55%),
        radial-gradient(800px 480px at 110% 10%, rgba(232, 197, 71, 0.16), transparent 50%),
        radial-gradient(700px 400px at 50% 120%, rgba(61, 220, 151, 0.08), transparent 50%),
        var(--paper);
    }
    body::before {
      content: "";
      position: fixed; inset: 0;
      background-image:
        linear-gradient(rgba(255,255,255,0.035) 1px, transparent 1px),
        linear-gradient(90deg, rgba(255,255,255,0.035) 1px, transparent 1px);
      background-size: 56px 56px;
      mask-image: radial-gradient(ellipse at center, black 30%, transparent 75%);
      pointer-events: none;
    }
    ::selection { background: var(--brand); color: var(--paper); }
    @keyframes rise-login {
      from { opacity: 0; transform: translateY(10px); }
      to   { opacity: 1; transform: none; }
    }
    .stage {
      width: 100%;
      max-width: 420px;
      position: relative;
      z-index: 1;
      animation: rise-login 0.45s cubic-bezier(0.16, 1, 0.3, 1) both;
    }

    header.brand {
      display: flex;
      align-items: center;
      gap: 12px;
      justify-content: center;
      margin-bottom: 32px;
    }
    .brand-mark {
      width: 44px; height: 44px;
      display: flex; align-items: center; justify-content: center;
      background: var(--brand);
      color: var(--paper);
      border-radius: 12px;
      font-family: var(--display);
      font-size: 22px;
      font-weight: 700;
      line-height: 1;
      box-shadow: 0 8px 24px -8px rgba(232, 197, 71, 0.7);
    }
    .brand-text .name {
      font-family: var(--display);
      font-size: 20px;
      font-weight: 600;
      letter-spacing: -0.02em;
    }
    .brand-text .sub {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.12em;
    }

    .card {
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 20px;
      padding: 40px 36px 32px;
      backdrop-filter: blur(20px);
      box-shadow: 0 24px 80px -24px rgba(0, 0, 0, 0.6);
    }

    h1.title {
      font-family: var(--display);
      font-size: 32px;
      font-weight: 600;
      line-height: 1.15;
      letter-spacing: -0.03em;
      margin: 0 0 8px;
    }
    h1.title em { font-style: normal; color: var(--brand); }
    .subtitle {
      color: var(--graphite);
      font-size: 14.5px;
      margin: 0 0 28px;
    }

    form { display: flex; flex-direction: column; gap: 18px; }
    .field { display: flex; flex-direction: column; gap: 6px; }
    .field > label {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.1em;
      font-weight: 600;
    }
    .field input {
      appearance: none;
      background: rgba(7, 11, 20, 0.55);
      border: 1px solid var(--border-strong);
      border-radius: 10px;
      padding: 12px 14px;
      font: inherit;
      font-size: 15px;
      color: var(--ink);
      transition: border-color 0.15s, box-shadow 0.15s, background 0.15s;
    }
    .field input:hover { border-color: rgba(232, 197, 71, 0.4); }
    .field input:focus {
      outline: none;
      border-color: var(--brand);
      background: rgba(7, 11, 20, 0.8);
      box-shadow: 0 0 0 3px var(--brand-soft);
    }

    .remember {
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 13px;
      color: var(--graphite);
      cursor: pointer;
      user-select: none;
    }
    .remember input {
      width: 16px; height: 16px;
      accent-color: var(--brand);
      cursor: pointer;
    }

    button.submit {
      background: var(--brand);
      color: var(--paper);
      border: none;
      border-radius: 10px;
      padding: 14px 20px;
      font: inherit;
      font-size: 15px;
      font-weight: 700;
      cursor: pointer;
      margin-top: 6px;
      box-shadow: 0 10px 28px -10px rgba(232, 197, 71, 0.7);
      transition: background 0.15s, transform 0.15s, box-shadow 0.15s;
    }
    button.submit:hover {
      background: var(--brand-hover);
      transform: translateY(-1px);
      box-shadow: 0 14px 32px -10px rgba(232, 197, 71, 0.85);
    }
    button.submit:active { transform: translateY(0); }
    button.submit:focus-visible {
      outline: none;
      box-shadow: 0 0 0 3px var(--brand-soft), 0 0 0 5px var(--brand);
    }

    .error {
      background: var(--rejected-soft);
      border: 1px solid rgba(240, 113, 103, 0.35);
      color: var(--rejected);
      border-radius: 10px;
      padding: 10px 14px;
      font-size: 13px;
      margin: 0 0 4px;
    }
    .footnote {
      text-align: center;
      color: var(--muted);
      font-size: 12px;
      margin-top: 20px;
    }
  </style>
</head>
<body>
  <div class="stage">
    <header class="brand">
      <div class="brand-mark">S</div>
      <div class="brand-text">
        <div class="name">Sourcing</div>
        <div class="sub">Amazon FBM discovery</div>
      </div>
    </header>

    <div class="card">
      <h1 class="title">Welcome <em>back</em>.</h1>
      <p class="subtitle">Sign in to continue.</p>

      {% if error %}<div class="error">{{ error }}</div>{% endif %}

      <form method="post" autocomplete="on">
        <div class="field">
          <label for="username">Username</label>
          <input id="username" name="username" type="text"
                 autocomplete="username" required autofocus>
        </div>
        <div class="field">
          <label for="password">Password</label>
          <input id="password" name="password" type="password"
                 autocomplete="current-password" required>
        </div>
        <label class="remember">
          <input type="checkbox" name="remember" checked>
          Keep me signed in for 30 days
        </label>
        <button type="submit" class="submit">Sign in</button>
      </form>
    </div>

    <p class="footnote">Contact your admin if you can't sign in.</p>
  </div>
</body>
</html>
"""


INDEX_HTML = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Sourcing</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
  <style>
    :root {
      --paper:        #070B14;
      --paper-warm:   #0E1524;
      --card:         rgba(18, 26, 42, 0.86);
      --border:       rgba(255, 255, 255, 0.08);
      --border-strong:rgba(255, 255, 255, 0.16);
      --ink:          #F2F5FA;
      --graphite:     #B7C0D1;
      --muted:        #7D8AA3;
      --brand:        #E8C547;
      --brand-hover:  #F3D56A;
      --brand-soft:   rgba(232, 197, 71, 0.16);
      --approved:     #3DDC97;
      --approved-soft:rgba(61, 220, 151, 0.12);
      --review:       #F0B429;
      --review-soft:  rgba(240, 180, 41, 0.12);
      --rejected:     #F07167;
      --rejected-soft:rgba(240, 113, 103, 0.14);
      --shadow-sm:  0 1px 0 rgba(255,255,255,0.04);
      --shadow-md:  0 18px 50px -24px rgba(0,0,0,0.7);
      --shadow-lg:  0 28px 80px -28px rgba(0,0,0,0.75);
      --sans:     "Plus Jakarta Sans", Inter, system-ui, sans-serif;
      --serif:    "Outfit", "Plus Jakarta Sans", system-ui, sans-serif;
      --display:  "Outfit", "Plus Jakarta Sans", system-ui, sans-serif;
      --mono:     ui-monospace, "SF Mono", "JetBrains Mono", Menlo, Consolas, monospace;
    }

    * { box-sizing: border-box; }
    html, body { margin: 0; padding: 0; }
    html { scroll-behavior: smooth; }
    body {
      font-family: var(--sans);
      color: var(--ink);
      line-height: 1.5;
      color-scheme: dark;
      -webkit-font-smoothing: antialiased;
      background:
        radial-gradient(1100px 560px at 8% -18%, rgba(99, 102, 241, 0.26), transparent 55%),
        radial-gradient(900px 500px at 100% 0%, rgba(232, 197, 71, 0.14), transparent 48%),
        radial-gradient(800px 420px at 50% 110%, rgba(61, 220, 151, 0.08), transparent 50%),
        var(--paper);
    }
    body::before {
      content: "";
      position: fixed; inset: 0;
      background-image:
        linear-gradient(rgba(255,255,255,0.03) 1px, transparent 1px),
        linear-gradient(90deg, rgba(255,255,255,0.03) 1px, transparent 1px);
      background-size: 56px 56px;
      mask-image: radial-gradient(ellipse at 50% 0%, black 20%, transparent 80%);
      pointer-events: none;
    }
    ::selection { background: var(--brand); color: var(--paper); }
    .serif { font-family: var(--display); }

    @keyframes rise {
      from { opacity: 0; transform: translateY(10px); }
      to   { opacity: 1; transform: none; }
    }
    .run-card,
    .progress-card.visible,
    .results-card.visible { animation: rise 0.4s cubic-bezier(0.16, 1, 0.3, 1) both; }

    .container {
      position: relative;
      z-index: 1;
      max-width: 860px;
      margin: 0 auto;
      padding: 36px 24px 96px;
    }

    header.brand {
      display: flex;
      align-items: center;
      gap: 12px;
      padding: 4px 0 28px;
      border-bottom: 1px solid var(--border);
      margin-bottom: 40px;
    }
    .brand-mark {
      width: 40px; height: 40px;
      display: flex; align-items: center; justify-content: center;
      background: var(--brand);
      color: var(--paper);
      border-radius: 11px;
      font-family: var(--display);
      font-size: 20px;
      font-weight: 700;
      line-height: 1;
      box-shadow: 0 8px 20px -8px rgba(232, 197, 71, 0.7);
    }
    .brand-text .name {
      font-family: var(--display);
      font-size: 18px;
      font-weight: 600;
      letter-spacing: -0.02em;
    }
    .brand-text .sub {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.12em;
    }
    .user-menu {
      margin-left: auto;
      display: flex;
      align-items: center;
      gap: 14px;
    }
    .user-name { font-size: 13px; color: var(--graphite); font-weight: 500; }
    .signout {
      font-size: 12px;
      color: var(--muted);
      text-decoration: none;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      font-weight: 600;
      padding: 7px 14px;
      border: 1px solid var(--border-strong);
      border-radius: 999px;
      transition: color 0.12s, border-color 0.12s, background 0.12s;
    }
    .signout:hover {
      color: var(--brand);
      border-color: var(--brand);
      background: var(--brand-soft);
    }

    .hero { margin-bottom: 36px; }
    .hero h1 {
      font-family: var(--display);
      font-size: 46px;
      font-weight: 600;
      line-height: 1.08;
      letter-spacing: -0.035em;
      margin: 0 0 14px;
      color: var(--ink);
    }
    @media (max-width: 560px) { .hero h1 { font-size: 32px; } }
    .hero h1 em { font-style: normal; color: var(--brand); }
    .hero p {
      color: var(--graphite);
      font-size: 15.5px;
      margin: 0;
      max-width: 560px;
      line-height: 1.6;
    }

    .run-card, .progress-card, .results-header, .approved-card, .reasons-card {
      background: var(--card);
      border: 1px solid var(--border);
      backdrop-filter: blur(18px);
      box-shadow: var(--shadow-md);
    }
    .run-card { border-radius: 20px; overflow: hidden; }
    .run-body { padding: 36px; }
    @media (max-width: 560px) { .run-body { padding: 24px; } }

    .form-section { margin-bottom: 32px; }
    .form-section:last-of-type { margin-bottom: 0; }
    .section-eyebrow {
      font-size: 10px;
      color: var(--brand);
      text-transform: uppercase;
      letter-spacing: 0.16em;
      font-weight: 700;
      margin-bottom: 6px;
    }
    .section-heading {
      font-family: var(--display);
      font-size: 24px;
      font-weight: 600;
      color: var(--ink);
      margin: 0 0 20px;
      line-height: 1.2;
      letter-spacing: -0.03em;
    }
    .section-heading em { font-style: normal; color: var(--brand); }
    .form-divider {
      height: 1px;
      background: var(--border);
      margin: 32px 0;
      border: none;
    }

    .fieldset {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 24px;
      margin-bottom: 20px;
    }
    .fieldset:last-child { margin-bottom: 0; }
    @media (max-width: 560px) { .fieldset { grid-template-columns: 1fr; } }

    .field { display: flex; flex-direction: column; gap: 8px; }
    .field > label {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.1em;
      font-weight: 600;
    }
    .field > label .opt {
      text-transform: none;
      letter-spacing: 0;
      font-weight: 400;
      opacity: 0.75;
      margin-left: 4px;
    }
    .field select, .field input[type="number"], .field input[type="text"] {
      appearance: none;
      -webkit-appearance: none;
      background: rgba(7, 11, 20, 0.55);
      border: 1px solid var(--border-strong);
      border-radius: 10px;
      padding: 12px 14px;
      font-size: 15px;
      color: var(--ink);
      font-family: inherit;
      transition: border-color 0.15s, box-shadow 0.15s, background 0.15s;
    }
    .field select:hover, .field input:hover { border-color: rgba(232, 197, 71, 0.4); }
    .field-hint { font-size: 11px; color: var(--muted); line-height: 1.45; }
    .field select {
      background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8' fill='none'%3E%3Cpath d='M1 1.5L6 6.5L11 1.5' stroke='%23B7C0D1' stroke-width='1.4'/%3E%3C/svg%3E");
      background-repeat: no-repeat;
      background-position: right 14px center;
      padding-right: 40px;
    }
    .field input:focus, .field select:focus {
      outline: none;
      border-color: var(--brand);
      background: rgba(7, 11, 20, 0.8);
      box-shadow: 0 0 0 3px var(--brand-soft);
    }

    .preset-row {
      display: grid;
      grid-template-columns: 1fr 1fr 1fr;
      gap: 12px;
    }
    @media (max-width: 640px) { .preset-row { grid-template-columns: 1fr; } }
    .preset {
      position: relative;
      background: rgba(7, 11, 20, 0.4);
      border: 1px solid var(--border-strong);
      border-radius: 14px;
      padding: 18px 16px;
      text-align: left;
      cursor: pointer;
      font-family: inherit;
      color: var(--ink);
      transition: border-color 0.18s, background 0.18s, transform 0.18s, box-shadow 0.18s;
      display: flex;
      flex-direction: column;
      gap: 4px;
    }
    .preset:hover:not(.selected) {
      border-color: rgba(232, 197, 71, 0.45);
      transform: translateY(-1px);
      background: rgba(7, 11, 20, 0.65);
    }
    .preset.selected {
      border-color: var(--brand);
      background: var(--brand-soft);
      box-shadow: 0 0 0 3px rgba(232, 197, 71, 0.12),
                  inset 3px 0 0 var(--brand);
    }
    .preset:focus-visible {
      outline: none;
      box-shadow: 0 0 0 3px var(--brand-soft);
    }
    .preset-label { font-weight: 700; font-size: 14px; }
    .preset-sub {
      font-size: 11px;
      color: var(--brand);
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }
    .preset-desc {
      font-size: 12px;
      color: var(--graphite);
      line-height: 1.4;
      margin-top: 6px;
    }

    details.advanced {
      border-top: 1px solid var(--border);
      margin: 32px -36px 0;
      padding: 0 36px;
    }
    @media (max-width: 560px) {
      details.advanced { margin: 32px -24px 0; padding: 0 24px; }
    }
    details.advanced summary {
      cursor: pointer;
      padding: 20px 0;
      font-size: 13px;
      font-weight: 600;
      color: var(--brand);
      list-style: none;
      display: flex;
      align-items: center;
      gap: 10px;
      user-select: none;
    }
    details.advanced summary::-webkit-details-marker { display: none; }
    details.advanced summary::before {
      content: "+";
      display: inline-block;
      width: 22px; height: 22px;
      background: var(--brand);
      color: var(--paper);
      border-radius: 7px;
      text-align: center;
      line-height: 21px;
      font-weight: 700;
      font-size: 15px;
    }
    details.advanced[open] summary::before { content: "\2212"; }

    .advanced-body {
      padding-bottom: 24px;
      display: grid;
      grid-template-columns: 1fr 1fr 1fr;
      gap: 20px;
    }
    @media (max-width: 640px) { .advanced-body { grid-template-columns: 1fr 1fr; } }
    .adv-group { display: flex; flex-direction: column; gap: 6px; }
    .adv-group > label {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.06em;
      font-weight: 600;
    }
    .adv-group input[type="number"] {
      background: rgba(7, 11, 20, 0.55);
      border: 1px solid var(--border-strong);
      border-radius: 10px;
      padding: 9px 11px;
      font-size: 14px;
      color: var(--ink);
      font-family: inherit;
      transition: border-color 0.15s, box-shadow 0.15s;
    }
    .adv-group input:hover { border-color: rgba(232, 197, 71, 0.4); }
    .adv-group input:focus {
      outline: none;
      border-color: var(--brand);
      box-shadow: 0 0 0 3px var(--brand-soft);
    }
    .adv-check {
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 13px;
      color: var(--graphite);
      cursor: pointer;
      padding: 6px 0;
    }
    .adv-check input { margin: 0; accent-color: var(--brand); }

    .run-footer {
      background: rgba(7, 11, 20, 0.45);
      border-top: 1px solid var(--border);
      padding: 22px 36px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 20px;
    }
    @media (max-width: 560px) {
      .run-footer {
        padding: 20px 24px;
        flex-direction: column;
        align-items: stretch;
      }
      .run-footer .btn-run { justify-content: center; }
    }
    .run-footer .footnote {
      font-size: 12px;
      color: var(--muted);
      max-width: 360px;
      line-height: 1.5;
    }
    .btn-run {
      position: relative;
      background: var(--brand);
      color: var(--paper);
      border: none;
      padding: 14px 28px 14px 32px;
      border-radius: 12px;
      font-size: 15px;
      font-weight: 700;
      cursor: pointer;
      font-family: inherit;
      display: inline-flex;
      align-items: center;
      gap: 12px;
      box-shadow: 0 12px 28px -10px rgba(232, 197, 71, 0.7);
      transition: background 0.15s, transform 0.15s, box-shadow 0.15s;
    }
    .btn-run .arrow {
      font-family: var(--display);
      font-size: 18px;
      line-height: 1;
      transition: transform 0.2s;
    }
    .btn-run:hover:not(:disabled) {
      background: var(--brand-hover);
      transform: translateY(-1px);
      box-shadow: 0 16px 36px -10px rgba(232, 197, 71, 0.85);
    }
    .btn-run:hover:not(:disabled) .arrow { transform: translateX(3px); }
    .btn-run:active:not(:disabled) { transform: translateY(0); }
    .btn-run:focus-visible {
      outline: none;
      box-shadow: 0 0 0 3px var(--brand-soft), 0 0 0 5px var(--brand);
    }
    .btn-run:disabled { cursor: not-allowed; opacity: 0.65; }
    .btn-run .spinner {
      display: none;
      width: 14px; height: 14px;
      border: 2px solid rgba(7, 11, 20, 0.25);
      border-top-color: var(--paper);
      border-radius: 50%;
      animation: spin 0.8s linear infinite;
    }
    .btn-run.loading .arrow { display: none; }
    .btn-run.loading .spinner { display: inline-block; }
    @keyframes spin { to { transform: rotate(360deg); } }

    .progress-card {
      display: none;
      border-radius: 20px;
      padding: 36px;
      margin-top: 20px;
    }
    @media (max-width: 560px) { .progress-card { padding: 24px; } }
    .progress-card.visible { display: block; }

    .stages {
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      margin-bottom: 28px;
    }
    @media (max-width: 560px) { .stages { grid-template-columns: 1fr 1fr; } }
    .stage {
      position: relative;
      padding: 14px;
      border: 1px solid var(--border);
      border-radius: 14px;
      background: rgba(7, 11, 20, 0.45);
      display: flex;
      flex-direction: column;
      gap: 6px;
      opacity: 0.5;
      transition: opacity 0.3s, background 0.3s, border-color 0.3s, box-shadow 0.3s;
    }
    .stage.done {
      opacity: 1;
      background: var(--approved-soft);
      border-color: rgba(61, 220, 151, 0.4);
    }
    .stage.active {
      opacity: 1;
      background: rgba(18, 26, 42, 0.95);
      border-color: var(--brand);
      box-shadow: 0 0 0 3px var(--brand-soft);
    }
    .stage.active::after {
      content: "";
      position: absolute;
      top: 10px; right: 10px;
      width: 8px; height: 8px;
      border-radius: 50%;
      background: var(--brand);
      animation: pulse 1.8s ease-in-out infinite;
    }
    @keyframes pulse {
      0%   { box-shadow: 0 0 0 0 rgba(232, 197, 71, 0.45); }
      70%  { box-shadow: 0 0 0 8px rgba(232, 197, 71, 0); }
      100% { box-shadow: 0 0 0 0 rgba(232, 197, 71, 0); }
    }
    .stage-num {
      font-family: var(--display);
      font-size: 22px;
      font-weight: 600;
      color: var(--brand);
      line-height: 1;
    }
    .stage.done .stage-num { color: var(--approved); font-size: 0; }
    .stage.done .stage-num::after {
      content: "\2713";
      font-family: var(--sans);
      font-size: 18px;
      font-weight: 700;
      line-height: 1;
      color: var(--approved);
    }
    .stage-label { font-size: 12px; font-weight: 700; color: var(--ink); }
    .stage-sub { font-size: 11px; color: var(--muted); line-height: 1.3; }

    .progress-bar-wrap { margin: 8px 0 16px; display: none; }
    .progress-bar-wrap.visible { display: block; }
    .progress-header {
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      margin-bottom: 8px;
      font-size: 13px;
    }
    .progress-header .label { color: var(--ink); font-weight: 700; }
    .progress-header .counter { color: var(--muted); font-variant-numeric: tabular-nums; }
    .progress-bar {
      width: 100%;
      height: 10px;
      background: rgba(7, 11, 20, 0.7);
      border-radius: 999px;
      overflow: hidden;
      border: 1px solid var(--border);
    }
    .progress-bar-fill {
      height: 100%;
      width: 0%;
      background: linear-gradient(90deg, #C9A227 0%, #E8C547 50%, #F3D56A 100%);
      border-radius: 999px;
      position: relative;
      overflow: hidden;
      transition: width 0.6s cubic-bezier(0.22, 1, 0.36, 1);
    }
    .progress-bar-fill::after {
      content: "";
      position: absolute; inset: 0;
      background-image: linear-gradient(
        45deg,
        rgba(255,255,255,0.18) 25%, transparent 25%,
        transparent 50%, rgba(255,255,255,0.18) 50%,
        rgba(255,255,255,0.18) 75%, transparent 75%
      );
      background-size: 16px 16px;
      animation: barber 1.5s linear infinite;
    }
    @keyframes barber {
      from { background-position: 0 0; }
      to   { background-position: 16px 0; }
    }
    .progress-detail {
      margin-top: 8px;
      font-size: 12px;
      color: var(--muted);
      display: flex;
      justify-content: space-between;
      gap: 12px;
      flex-wrap: wrap;
    }
    .progress-detail .activity {
      flex: 1; min-width: 0;
      white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }
    .progress-detail .eta {
      color: var(--brand);
      font-variant-numeric: tabular-nums;
      white-space: nowrap;
    }
    .progress-pills { display: flex; gap: 8px; margin-bottom: 16px; flex-wrap: wrap; }
    .pill {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 5px 11px;
      background: rgba(7, 11, 20, 0.55);
      border: 1px solid var(--border);
      border-radius: 999px;
      font-size: 11px;
      color: var(--muted);
      font-variant-numeric: tabular-nums;
    }
    .pill strong { color: var(--ink); font-weight: 700; }
    .pill.warn {
      background: rgba(240, 180, 41, 0.14);
      border-color: rgba(240, 180, 41, 0.4);
      color: #F0B429;
    }
    .pill.warn strong { color: #F3D56A; }

    .banner-warn {
      display: none;
      background: rgba(240, 180, 41, 0.12);
      border: 1px solid rgba(240, 180, 41, 0.35);
      border-left: 4px solid #F0B429;
      color: #F3D56A;
      padding: 12px 16px;
      border-radius: 12px;
      margin-bottom: 16px;
      font-size: 13px;
      line-height: 1.5;
    }
    .banner-warn.visible { display: block; }
    .banner-warn strong { color: #F3D56A; }

    .log-shell { position: relative; }
    .log-shell::before {
      content: "Live log";
      display: block;
      font-size: 10px;
      font-weight: 700;
      letter-spacing: 0.14em;
      text-transform: uppercase;
      color: var(--muted);
      padding: 0 0 8px;
    }
    .log-view {
      background: #05080F;
      color: #C5D0E0;
      font-family: var(--mono);
      font-size: 11.5px;
      padding: 18px 20px;
      border-radius: 12px;
      max-height: 220px;
      overflow-y: auto;
      white-space: pre-wrap;
      line-height: 1.7;
      border: 1px solid var(--border);
      scrollbar-width: thin;
      scrollbar-color: #33415C transparent;
    }
    .log-view::-webkit-scrollbar { width: 6px; }
    .log-view::-webkit-scrollbar-thumb { background: #33415C; border-radius: 3px; }

    .results-card { display: none; margin-top: 20px; }
    .results-card.visible { display: block; }
    .results-header {
      border-radius: 20px;
      padding: 36px;
      margin-bottom: 20px;
    }
    @media (max-width: 560px) { .results-header { padding: 24px; } }
    .results-header h2 {
      font-family: var(--display);
      font-size: 30px;
      font-weight: 600;
      margin: 0 0 24px;
      letter-spacing: -0.03em;
    }
    .results-header h2 em { font-style: normal; color: var(--brand); }
    .results-meta {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: -12px 0 20px;
    }
    .results-meta:empty { display: none; }

    .stat-row {
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 12px;
      margin-bottom: 24px;
    }
    @media (max-width: 560px) { .stat-row { grid-template-columns: 1fr 1fr; } }
    .stat {
      padding: 18px 16px;
      border-radius: 14px;
      border: 1px solid var(--border);
      background: rgba(7, 11, 20, 0.45);
    }
    .stat .num {
      font-family: var(--display);
      font-size: 40px;
      font-weight: 600;
      line-height: 1;
      letter-spacing: -0.03em;
      font-variant-numeric: tabular-nums;
    }
    .stat .lbl {
      font-size: 11px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.1em;
      margin-top: 10px;
      font-weight: 700;
    }
    .stat.approved { background: var(--approved-soft); border-color: rgba(61, 220, 151, 0.4); }
    .stat.approved .num { color: var(--approved); }
    .stat.review { background: var(--review-soft); border-color: rgba(240, 180, 41, 0.4); }
    .stat.review .num { color: var(--review); }
    .stat.rejected { background: var(--rejected-soft); border-color: rgba(240, 113, 103, 0.4); }
    .stat.rejected .num { color: var(--rejected); }
    .stat.total .num { color: var(--ink); }

    .download-cta {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 22px 26px;
      background: linear-gradient(135deg, #E8C547 0%, #C9A227 100%);
      color: var(--paper);
      border-radius: 14px;
      box-shadow: 0 14px 36px -12px rgba(232, 197, 71, 0.55);
    }
    .download-cta .txt { font-size: 13px; opacity: 0.85; line-height: 1.4; }
    .download-cta .btn-download {
      display: inline-flex;
      align-items: center;
      gap: 10px;
      background: var(--paper);
      color: var(--brand);
      padding: 12px 20px;
      border-radius: 10px;
      font-weight: 700;
      text-decoration: none;
      font-size: 14px;
      white-space: nowrap;
      transition: transform 0.15s, box-shadow 0.15s;
    }
    .download-cta .btn-download::before {
      content: "";
      display: inline-block;
      width: 15px; height: 15px;
      background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='15' height='15' viewBox='0 0 24 24' fill='none' stroke='%23E8C547' stroke-width='2.4' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4'/%3E%3Cpolyline points='7 10 12 15 17 10'/%3E%3Cline x1='12' y1='15' x2='12' y2='3'/%3E%3C/svg%3E");
      background-repeat: no-repeat;
      background-position: center;
    }
    .download-cta .btn-download:hover {
      transform: translateY(-1px);
      box-shadow: 0 8px 18px -8px rgba(0,0,0,0.45);
    }

    .section-title {
      font-family: var(--display);
      font-size: 20px;
      font-weight: 600;
      margin: 32px 0 8px;
      letter-spacing: -0.02em;
    }
    .section-hint { font-size: 12px; color: var(--muted); margin-bottom: 16px; }

    .approved-card, .reasons-card {
      border-radius: 16px;
      padding: 4px 24px;
    }
    .approved-item {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 20px;
      padding: 18px 12px;
      margin: 0 -12px;
      border-bottom: 1px solid var(--border);
      border-radius: 10px;
    }
    .approved-item:hover { background: rgba(232, 197, 71, 0.06); }
    .approved-item:last-child { border-bottom: none; }
    .approved-item .info { min-width: 0; }
    .approved-item .title {
      font-size: 14.5px;
      font-weight: 600;
      margin-bottom: 8px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .approved-item .meta {
      font-size: 11.5px;
      color: var(--muted);
      display: flex;
      gap: 14px;
      flex-wrap: wrap;
      align-items: center;
    }
    .approved-item .meta a {
      color: var(--paper);
      text-decoration: none;
      font-weight: 700;
      font-family: var(--mono);
      font-size: 11px;
      padding: 2px 7px;
      background: var(--brand);
      border-radius: 5px;
    }
    .approved-item .meta a:hover { background: var(--brand-hover); }
    .approved-item .profit {
      text-align: right;
      font-family: var(--display);
      font-size: 26px;
      font-weight: 600;
      color: var(--approved);
      line-height: 1;
      white-space: nowrap;
      font-variant-numeric: tabular-nums;
    }
    .approved-item .roi {
      font-size: 11px;
      color: var(--muted);
      text-align: right;
      margin-top: 8px;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      font-weight: 700;
    }

    .reason-row {
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding: 15px 0;
      border-bottom: 1px solid var(--border);
      font-size: 13px;
    }
    .reason-row:last-child { border-bottom: none; }
    .reason-row .name { color: var(--ink); }
    .reason-row .count {
      background: var(--rejected-soft);
      color: var(--rejected);
      padding: 4px 12px;
      border-radius: 8px;
      font-family: var(--mono);
      font-size: 12px;
      font-weight: 700;
    }

    .start-over {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      color: var(--muted);
      font-size: 13px;
      text-decoration: none;
      padding: 10px 18px;
      border: 1px solid var(--border-strong);
      border-radius: 999px;
      font-weight: 600;
      transition: color 0.15s, border-color 0.15s, background 0.15s;
    }
    .start-over::before { content: "\21BB"; font-size: 14px; }
    .start-over:hover {
      color: var(--brand);
      border-color: var(--brand);
      background: var(--brand-soft);
    }
    .start-over-wrap { display: flex; justify-content: center; margin-top: 32px; }

    .empty-state { text-align: center; padding: 32px 24px; color: var(--muted); }
    .empty-state h3 {
      font-family: var(--display);
      font-weight: 600;
      color: var(--ink);
      font-size: 20px;
      margin-bottom: 8px;
    }

    .error-box {
      background: var(--rejected-soft);
      border: 1px solid rgba(240, 113, 103, 0.4);
      border-radius: 12px;
      padding: 16px 20px;
      margin-bottom: 20px;
      color: var(--rejected);
      font-size: 13px;
      white-space: pre-wrap;
      font-family: var(--mono);
      max-height: 320px;
      overflow-y: auto;
    }
    .error-box strong {
      display: block;
      margin-bottom: 8px;
      font-family: var(--sans);
    }
  </style>
</head>
<body>

<div class="container">

  <header class="brand">
    <div class="brand-mark serif">S</div>
    <div class="brand-text">
      <div class="name">Sourcing</div>
      <div class="sub">Amazon FBM discovery</div>
    </div>
    {% if auth_enabled %}
    <div class="user-menu">
      <span class="user-name">{{ username }}</span>
      <a class="signout" href="/logout" title="Sign out">Sign out</a>
    </div>
    {% endif %}
  </header>

  <section class="hero">
    <h1>Find products worth <em>listing</em>.</h1>
    <p>Pick a supplier, choose how many products to analyze, and get back a filtered list with real profit numbers and Amazon competition data.</p>
  </section>

  <form id="runForm">
    <div class="run-card">
      <div class="run-body">

        <div class="form-section">
          <div class="section-eyebrow">Step 1</div>
          <h2 class="section-heading">Choose your <em>source</em>.</h2>

          <div class="fieldset">
            <div class="field">
              <label for="retailer">Supplier</label>
              <select id="retailer" name="retailer">
                {% for r in retailers %}
                <option value="{{ r.value }}">{{ r.label }} - {{ r.sub }}</option>
                {% endfor %}
              </select>
            </div>
            <div class="field">
              <label for="limit">Products to analyze</label>
              <input type="number" id="limit" name="limit" value="25" min="1" max="500">
            </div>
          </div>

          <div class="fieldset">
            <div class="field">
              <label for="brands">Focus on brands <span class="opt">(optional)</span></label>
              <input type="text" id="brands" placeholder='e.g., Now Foods, Jarrow Formulas'>
              <div class="field-hint">Comma-separated. Only products from these brands will be scraped.</div>
            </div>
            <div class="field">
              <label>&nbsp;</label>
              <div class="field-hint" style="padding-top: 8px;">
                For iHerb, the tool uses the brand listing with iHerb's own <em>In-stock</em> filter, then double-checks each product page.
              </div>
            </div>
          </div>
        </div>

        <hr class="form-divider">

        <div class="form-section">
          <div class="section-eyebrow">Step 2</div>
          <h2 class="section-heading">Pick your <em>filters</em>.</h2>

          <div class="preset-row">
            {% for key, p in presets.items() %}
            <button class="preset {% if key == 'balanced' %}selected{% endif %}"
                    data-preset="{{ key }}" type="button">
              <div class="preset-label">{{ p.label }}</div>
              <div class="preset-sub">{{ p.sub }}</div>
              <div class="preset-desc">{{ p.desc }}</div>
            </button>
            {% endfor %}
          </div>
        </div>

        <details class="advanced">
          <summary>Adjust individual filters</summary>
          <div class="advanced-body">

            <div class="adv-group">
              <label>Min profit ($)</label>
              <input type="number" id="min_profit" step="0.01" value="2">
            </div>
            <div class="adv-group">
              <label>Min margin (%)</label>
              <input type="number" id="min_margin" value="10">
            </div>
            <div class="adv-group">
              <label>Min ROI (%)</label>
              <input type="number" id="min_roi" value="15">
            </div>

            <div class="adv-group">
              <label>Min FBM sellers</label>
              <input type="number" id="min_fbm_sellers" value="4">
            </div>
            <div class="adv-group">
              <label>Min rating</label>
              <input type="number" id="min_rating" step="0.1" value="3.5">
            </div>
            <div class="adv-group">
              <label>Min reviews</label>
              <input type="number" id="min_reviews" value="0">
            </div>

            <div class="adv-group">
              <label>Min historical sellers</label>
              <input type="number" id="min_historical_sellers" value="0">
            </div>
            <div class="adv-group">
              <label>Max BSR (0 = none)</label>
              <input type="number" id="max_bsr" value="0">
            </div>
            <div class="adv-group">
              <label>Amazon fee (%)</label>
              <input type="number" id="fee_pct" value="15">
            </div>

            <div class="adv-check" style="grid-column: 1 / -1;">
              <input type="checkbox" id="require_rating">
              <label for="require_rating">Require rating (reject if no rating at all)</label>
            </div>
            <div class="adv-check" style="grid-column: 1 / -1;">
              <input type="checkbox" id="require_reviews">
              <label for="require_reviews">Require reviews (reject if no reviews at all)</label>
            </div>

          </div>
        </details>

      </div>

      <div class="run-footer">
        <div class="footnote">
          Out-of-stock supplier products are skipped automatically.
          Analysis takes roughly 20 seconds per product. Each product costs ~7 Keepa tokens.
        </div>
        <button type="submit" class="btn-run" id="runBtn">
          <span class="label">Start analysis</span>
          <span class="arrow" aria-hidden="true">&rarr;</span>
          <span class="spinner" aria-hidden="true"></span>
        </button>
      </div>
    </div>
  </form>

  <div class="progress-card" id="progressCard">
    <div class="stages">
      <div class="stage" data-stage="discovering">
        <div class="stage-num">1</div>
        <div class="stage-label">Discovering</div>
        <div class="stage-sub">Reading supplier catalog</div>
      </div>
      <div class="stage" data-stage="scraping">
        <div class="stage-num">2</div>
        <div class="stage-label">Scraping</div>
        <div class="stage-sub">Fetching product details</div>
      </div>
      <div class="stage" data-stage="matching">
        <div class="stage-num">3</div>
        <div class="stage-label">Matching</div>
        <div class="stage-sub">Looking up on Amazon</div>
      </div>
      <div class="stage" data-stage="finalizing">
        <div class="stage-num">4</div>
        <div class="stage-label">Filtering</div>
        <div class="stage-sub">Applying your rules</div>
      </div>
    </div>

    <div class="banner-warn" id="keepaExhaustedBanner">
      <strong>Keepa quota hit.</strong>
      Partial results were saved. Products still to check are marked
      <em>INCOMPLETE</em> in the Excel file — re-run in a few minutes once
      Keepa refills your token bucket.
    </div>

    <div class="progress-bar-wrap" id="progressBarWrap">
      <div class="progress-header">
        <span class="label" id="progressStageLabel">Working</span>
        <span class="counter" id="progressCounter"></span>
      </div>
      <div class="progress-bar">
        <div class="progress-bar-fill" id="progressBarFill"></div>
      </div>
      <div class="progress-detail">
        <span class="activity" id="progressActivity">&nbsp;</span>
        <span class="eta" id="progressEta"></span>
      </div>
    </div>

    <div class="progress-pills" id="progressPills"></div>

    <div class="log-shell">
      <div class="log-view" id="logView"></div>
    </div>
  </div>

  <div class="results-card" id="resultsCard">
    <div class="results-header">

      <div class="error-box" id="errorBox" style="display: none;">
        <strong>Something went wrong.</strong>
        <span id="errorMessage"></span>
      </div>

      <div class="banner-warn" id="resultsKeepaExhaustedBanner">
        <strong>Keepa quota was hit during this run.</strong>
        Some products couldn't be fully analysed and are marked
        <em>INCOMPLETE</em> in the spreadsheet. Re-run in a few minutes to
        finish them once your token bucket refills.
      </div>

      <h2>Analysis <em>complete</em></h2>

      <div class="results-meta" id="resultsMeta"></div>

      <div class="stat-row" id="statRow"></div>

      <div class="download-cta" id="downloadCta">
        <div class="txt">All data lives in the spreadsheet, including rejected products.</div>
        <a class="btn-download" id="downloadBtn" href="#" target="_blank">Download Excel</a>
      </div>
    </div>

    <div id="approvedSection" style="display: none;">
      <h3 class="section-title">Top approved products</h3>
      <p class="section-hint">Ranked by profit per unit. Full list in the Excel file.</p>
      <div class="approved-card" id="approvedList"></div>
    </div>

    <div id="reasonsSection" style="display: none;">
      <h3 class="section-title">Why others were skipped</h3>
      <p class="section-hint">Change filters and re-run to see if these come through.</p>
      <div class="reasons-card" id="reasonsList"></div>
    </div>

    <div id="emptyState" class="empty-state" style="display: none;">
      <h3>No approved products this run.</h3>
      <p>Try loosening your criteria or running a larger batch.</p>
    </div>

    <div class="start-over-wrap">
      <a href="#" class="start-over" onclick="location.reload(); return false;">Run another analysis</a>
    </div>
  </div>

</div>

<script>
const PRESETS = {{ presets|tojson }};

document.querySelectorAll(".preset").forEach(btn => {
  btn.addEventListener("click", () => {
    document.querySelectorAll(".preset").forEach(b => b.classList.remove("selected"));
    btn.classList.add("selected");
    applyPreset(btn.dataset.preset);
  });
});

function applyPreset(key) {
  const v = PRESETS[key].values;
  set("min_profit", v.min_profit);
  set("min_margin", Math.round(v.min_margin * 100));
  set("min_roi", Math.round(v.min_roi * 100));
  set("min_fbm_sellers", v.min_fbm_sellers);
  set("min_rating", v.min_rating);
  set("min_reviews", v.min_reviews);
  set("min_historical_sellers", v.min_historical_sellers);
  set("max_bsr", v.max_bsr);
  document.getElementById("require_rating").checked = v.require_rating;
  document.getElementById("require_reviews").checked = v.require_reviews;
}
function set(id, val) { document.getElementById(id).value = val; }

function buildConfig() {
  const val = id => parseFloat(document.getElementById(id).value) || 0;
  const chk = id => document.getElementById(id).checked;
  const maxBsr = val("max_bsr");
  return {
    fee_pct: val("fee_pct") / 100,
    min_profit: val("min_profit"),
    min_margin: val("min_margin") / 100,
    min_roi: val("min_roi") / 100,
    min_fbm_sellers: parseInt(val("min_fbm_sellers")) || 0,
    min_historical_sellers: parseInt(val("min_historical_sellers")) || 0,
    reject_if_amazon_on_listing: false,
    min_rating: val("min_rating"),
    require_rating: chk("require_rating"),
    min_reviews: parseInt(val("min_reviews")) || 0,
    require_reviews: chk("require_reviews"),
    max_bsr: maxBsr > 0 ? maxBsr : null,
  };
}

function setRunButtonState(state) {
  const btn = document.getElementById("runBtn");
  const label = btn.querySelector(".label");
  btn.classList.remove("loading");
  btn.disabled = false;
  if (state === "loading") {
    btn.disabled = true;
    btn.classList.add("loading");
    label.textContent = "Working";
  } else {
    label.textContent = "Start analysis";
  }
}

document.getElementById("runForm").addEventListener("submit", async (e) => {
  e.preventDefault();
  setRunButtonState("loading");

  document.getElementById("resultsCard").classList.remove("visible");
  document.getElementById("progressCard").classList.add("visible");
  document.getElementById("logView").textContent = "";
  document.querySelectorAll(".stage").forEach(s => s.classList.remove("done", "active"));
  document.getElementById("progressBarWrap").classList.remove("visible");
  document.getElementById("progressBarFill").style.width = "0%";
  document.getElementById("progressCounter").textContent = "";
  document.getElementById("progressActivity").textContent = "\u00A0";
  document.getElementById("progressEta").textContent = "";
  document.getElementById("progressPills").innerHTML = "";
  document.getElementById("keepaExhaustedBanner").classList.remove("visible");

  const config = buildConfig();

  const payload = {
    sitemap_url: document.getElementById("retailer").value,
    limit: parseInt(document.getElementById("limit").value),
    brands: document.getElementById("brands").value.trim(),
    random: true,
    no_robots: true,
    config: config,
  };

  try {
    const resp = await fetch("/api/run", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(payload),
    });
    const data = await resp.json();
    if (data.error) throw new Error(data.error);
    document.getElementById("progressCard").scrollIntoView({ behavior: "smooth", block: "nearest" });
    pollStatus(data.job_id);
  } catch (err) {
    showError(err.message);
    setRunButtonState("idle");
  }
});

const STAGE_ORDER = ["discovering", "scraping", "matching", "finalizing"];
const STAGE_LABELS = {
  discovering: "Discovering products",
  scraping:    "Scraping supplier",
  matching:    "Matching on Amazon",
  finalizing:  "Applying your rules",
};

function fmtEta(sec) {
  if (sec == null || sec < 0) return "";
  if (sec < 60) return "~" + sec + "s left";
  const m = Math.floor(sec / 60), s = sec % 60;
  if (m < 60) return "~" + m + "m " + (s ? s + "s" : "") + " left";
  const h = Math.floor(m / 60);
  return "~" + h + "h " + (m % 60) + "m left";
}

function renderProgressPills(data) {
  const pills = [];
  if (data.urls_discovered != null) {
    pills.push('<span class="pill">Discovered <strong>' +
      data.urls_discovered + '</strong> URLs</span>');
  }
  if (data.keepa_tokens_left != null) {
    const cls = data.keepa_tokens_left < 100 ? "pill warn" : "pill";
    pills.push('<span class="' + cls + '">Keepa tokens <strong>' +
      data.keepa_tokens_left + '</strong></span>');
  }
  document.getElementById("progressPills").innerHTML = pills.join("");
}

async function pollStatus(jobId) {
  try {
    const resp = await fetch("/api/status/" + jobId);
    const data = await resp.json();

    document.getElementById("logView").textContent = data.log_tail.join("\n");
    const logEl = document.getElementById("logView");
    logEl.scrollTop = logEl.scrollHeight;

    // Stage highlighting
    if (data.stage && STAGE_ORDER.includes(data.stage)) {
      const idx = STAGE_ORDER.indexOf(data.stage);
      STAGE_ORDER.forEach((s, i) => {
        const el = document.querySelector('.stage[data-stage="' + s + '"]');
        el.classList.remove("active", "done");
        if (i < idx) el.classList.add("done");
        else if (i === idx) el.classList.add("active");
      });
    }

    // Per-item progress bar
    const barWrap = document.getElementById("progressBarWrap");
    if (data.current != null && data.total != null && data.total > 0) {
      barWrap.classList.add("visible");
      const pct = Math.min(100, Math.round(100 * data.current / data.total));
      document.getElementById("progressBarFill").style.width = pct + "%";
      document.getElementById("progressCounter").textContent =
        data.current + " / " + data.total;
      document.getElementById("progressStageLabel").textContent =
        STAGE_LABELS[data.stage] || "Working";
      document.getElementById("progressEta").textContent =
        fmtEta(data.eta_seconds);
    }
    if (data.current_activity) {
      document.getElementById("progressActivity").textContent =
        data.current_activity;
    }

    renderProgressPills(data);

    if (data.keepa_exhausted) {
      document.getElementById("keepaExhaustedBanner").classList.add("visible");
    }

    if (data.status === "done") {
      STAGE_ORDER.forEach(s => {
        const el = document.querySelector('.stage[data-stage="' + s + '"]');
        el.classList.remove("active");
        el.classList.add("done");
      });
      document.getElementById("progressBarFill").style.width = "100%";
      document.getElementById("progressEta").textContent = "";
      setTimeout(() => showResults(jobId, data.summary, data), 400);
    } else if (data.status === "error") {
      showError(data.error || "Unknown error");
    } else {
      setTimeout(() => pollStatus(jobId), 1500);
    }
  } catch (err) {
    showError("Lost connection: " + err.message);
  }
}

function showResults(jobId, summary, status) {
  document.getElementById("progressCard").classList.remove("visible");
  const results = document.getElementById("resultsCard");
  results.classList.add("visible");

  setRunButtonState("idle");

  document.getElementById("errorBox").style.display = "none";
  document.getElementById("downloadCta").style.display = "";

  // Persist Keepa signals into the results header so the user can plan the
  // next run (progress card is hidden at this point, so the token pill and
  // exhausted banner over there are no longer visible).
  const meta = document.getElementById("resultsMeta");
  meta.innerHTML = "";
  if (status && status.keepa_tokens_left != null) {
    const cls = status.keepa_tokens_left < 100 ? "pill warn" : "pill";
    meta.insertAdjacentHTML("beforeend",
      '<span class="' + cls + '">Keepa tokens remaining <strong>' +
      status.keepa_tokens_left + '</strong></span>');
  }
  const exhaustedBanner = document.getElementById("resultsKeepaExhaustedBanner");
  if (status && status.keepa_exhausted) {
    exhaustedBanner.classList.add("visible");
  } else {
    exhaustedBanner.classList.remove("visible");
  }

  const byStatus = summary.by_status || {};
  const approved = byStatus["APPROVED"] || 0;
  const review = byStatus["REVIEW"] || 0;
  const rejected = byStatus["Rejected"] || 0;
  const total = summary.total || 0;

  const statRow = document.getElementById("statRow");
  statRow.innerHTML = "";
  addStat(statRow, approved, "Approved", "approved");
  if (review > 0) addStat(statRow, review, "Review", "review");
  addStat(statRow, rejected, "Skipped", "rejected");
  addStat(statRow, total, "Analyzed", "total");

  const approvedList = document.getElementById("approvedList");
  approvedList.innerHTML = "";
  if (summary.top_approved && summary.top_approved.length) {
    document.getElementById("approvedSection").style.display = "";
    document.getElementById("emptyState").style.display = "none";
    summary.top_approved.forEach(p => {
      const row = document.createElement("div");
      row.className = "approved-item";
      const asinLink = p.asin
        ? '<a href="https://www.amazon.com/dp/' + p.asin + '" target="_blank">' + p.asin + '</a>'
        : "";
      const supplierCost = p.supplier_cost ? "$" + p.supplier_cost.toFixed(2) : "";
      const amazonPrice = p.amazon_price ? "$" + p.amazon_price.toFixed(2) : "";
      const priceInfo = supplierCost && amazonPrice
        ? supplierCost + " to " + amazonPrice : "";
      const fbmInfo = p.fbm ? p.fbm + " FBM sellers" : "";
      const roiPct = p.roi ? (p.roi * 100).toFixed(0) + "% ROI" : "";
      row.innerHTML =
        '<div class="info">' +
          '<div class="title">' + escapeHtml(p.title || "Untitled") + '</div>' +
          '<div class="meta">' +
            asinLink +
            (priceInfo ? '<span>' + priceInfo + '</span>' : '') +
            (fbmInfo ? '<span>' + fbmInfo + '</span>' : '') +
          '</div>' +
        '</div>' +
        '<div>' +
          '<div class="profit">$' + (p.profit || 0).toFixed(2) + '</div>' +
          '<div class="roi">' + roiPct + '</div>' +
        '</div>';
      approvedList.appendChild(row);
    });
  } else {
    document.getElementById("approvedSection").style.display = "none";
    document.getElementById("emptyState").style.display = "";
  }

  const reasonsList = document.getElementById("reasonsList");
  reasonsList.innerHTML = "";
  if (summary.reject_reasons && summary.reject_reasons.length) {
    document.getElementById("reasonsSection").style.display = "";
    summary.reject_reasons.forEach(pair => {
      const reason = pair[0];
      const count = pair[1];
      const row = document.createElement("div");
      row.className = "reason-row";
      row.innerHTML = '<span class="name">' + escapeHtml(reason) + '</span>' +
                      '<span class="count">' + count + '</span>';
      reasonsList.appendChild(row);
    });
  } else {
    document.getElementById("reasonsSection").style.display = "none";
  }

  document.getElementById("downloadBtn").href = "/api/download/" + jobId;
}

function showError(msg) {
  document.getElementById("progressCard").classList.remove("visible");
  const results = document.getElementById("resultsCard");
  results.classList.add("visible");

  document.getElementById("errorMessage").textContent = msg;
  document.getElementById("errorBox").style.display = "";
  document.getElementById("statRow").innerHTML = "";
  document.getElementById("resultsMeta").innerHTML = "";
  document.getElementById("resultsKeepaExhaustedBanner").classList.remove("visible");
  document.getElementById("downloadCta").style.display = "none";
  document.getElementById("approvedSection").style.display = "none";
  document.getElementById("reasonsSection").style.display = "none";
  document.getElementById("emptyState").style.display = "none";

  setRunButtonState("idle");
}

function addStat(container, num, label, cls) {
  const d = document.createElement("div");
  d.className = "stat " + cls;
  d.innerHTML = '<div class="num">' + num + '</div><div class="lbl">' + label + '</div>';
  container.appendChild(d);
}

function escapeHtml(s) {
  return String(s || "").replace(/[&<>"']/g,
    c => ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[c]));
}

applyPreset("balanced");
</script>

</body>
</html>
"""


# ---------------------------------------------------------------- #
# Entrypoint                                                       #
# ---------------------------------------------------------------- #

def main() -> None:
    port = int(os.environ.get("PORT", 5000))
    url = f"http://127.0.0.1:{port}"
    print("=" * 60)
    print(f"  Sourcing UI running at:  {url}")
    print("=" * 60)
    print("  Press Ctrl+C to stop.")
    print()

    try:
        import webbrowser
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    except Exception:
        pass

    app.run(host="127.0.0.1", port=port, debug=False)


if __name__ == "__main__":
    main()
