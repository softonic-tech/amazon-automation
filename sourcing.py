"""
Sourcing Analysis
-----------------
Combines Zoro product data + Amazon comparison + Amazon product page details
into an FBA sourcing worksheet matching the demo template.

Data model:
  SourcingRow = one Zoro product + one candidate Amazon match + P&L + decision

Fields marked NEEDS_KEEPA are placeholders — scraping cannot reliably provide
them. Wire in Keepa or a paid API to fill these:
  * FBM Sellers (count of Fulfilled-by-Merchant sellers on the listing)
  * BSR (Best Sellers Rank — sometimes visible, often not)

Sourcing rules (matches the demo):
  * Supplier out of stock     → REJECT (skipped at hunt time too)
  * Amazon on listing         → REJECT
  * FBM sellers < min_fbm     → REJECT (requires Keepa)
  * Profit $ <= 0             → REJECT (unprofitable)
  * Margin % < min_margin     → REJECT (optional)
  * ROI % < min_roi           → REJECT (optional)
  * else                      → APPROVED
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any

from amazon import AmazonProductDetails, AmazonResult, AmazonSearcher, compare_zoro_to_amazon

log = logging.getLogger("scraper.sourcing")

NEEDS_KEEPA = "NEEDS_KEEPA"

# schema.org Offer.availability values that mean "cannot buy right now".
# Matching is case-insensitive and ignores punctuation / URL prefixes
# (e.g. "https://schema.org/OutOfStock" → "outofstock").
_OUT_OF_STOCK_TOKENS = frozenset({
            "outofstock",
            "outofstocketa",
            "soldout",
    "discontinued",
    "backorder",
    "backordered",
    "preorder",
    "presale",
    "unavailable",
    "outofstockonline",
    "temporarilyoutofstock",
})


def is_out_of_stock(availability: str | None) -> bool:
    """True only when availability is known to be out of stock.

    Unknown / missing availability is treated as *not* out of stock so we
    don't drop products just because a page omitted the field.
    """
    if not availability:
        return False
    raw = str(availability).strip()
    if not raw:
        return False
    low = raw.lower()
    if "out of stock" in low or "sold out" in low or "not available" in low:
        return True
    token = re.sub(r"[^a-z0-9]", "", low.split("/")[-1])
    if token in _OUT_OF_STOCK_TOKENS:
        return True
    return "outofstock" in token or token.endswith("soldout")


@dataclass
class SourcingConfig:
    # --- Financial ---
    fee_pct: float = 0.15          # Amazon referral fee (typically 15%)
    extra_cost: float = 0.0        # per-unit shipping/prep buffer
    min_supplier_price: float = 0.0  # skip supplier products below this price

    # --- Profitability filters ---
    min_profit: float = 0.01       # min $ profit; 0.01 = "any positive profit"
    min_margin: float = 0.0        # min margin, e.g. 0.10 for 10%
    min_roi: float = 0.0           # min ROI, e.g. 0.15 for 15%

    # --- Competition filters ---
    min_fbm_sellers: int = 4       # client's ≥4 FBM rule
    max_fbm_sellers: int | None = None       # None = unlimited (avoid over-crowded listings if set)
    reject_if_amazon_on_listing: bool = True
    min_historical_sellers: int = 0          # 0 = don't apply; higher = only mature listings

    # --- Product quality filters ---
    min_rating: float = 3.5        # client's ≥3.5 rule
    require_rating: bool = False    # if True, reject "no rating" (unknown rating fails)
    min_reviews: int = 0            # 0 = don't apply
    require_reviews: bool = False   # if True, reject "no reviews"
    require_in_stock: bool = True   # skip/reject supplier out-of-stock products

    # --- Velocity filters ---
    max_bsr: int | None = None              # None = don't apply; e.g. 100000 = mainstream only

    # --- Currency handling ---
    # For non-USD suppliers (pk.iherb.com serves PKR). Rate = "1 USD equals this many source units."
    # Example: for PKR, use ~280. For USD supplier data, leave at 1.0.
    supplier_currency: str = "USD"
    supplier_to_usd_rate: float = 1.0

    # --- Meta ---
    config_source: str = "defaults"  # populated when loaded from file

    @classmethod
    def from_file(cls, path: str | Path) -> "SourcingConfig":
        """
        Load config from a JSON file. Unknown fields are ignored (forward-compatible).
        Missing fields keep their defaults.
        """
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Config file not found: {p}")
        data = json.loads(p.read_text(encoding="utf-8"))
        # Filter to only known fields — allows commented-out or extra keys
        valid_fields = {f for f in cls.__dataclass_fields__ if f != "config_source"}
        clean = {k: v for k, v in data.items() if k in valid_fields}
        cfg = cls(**clean)
        cfg.config_source = str(p)
        return cfg

    def save_to_file(self, path: str | Path) -> None:
        """Save current config to a JSON file (for the client to edit later)."""
        from dataclasses import asdict
        data = asdict(self)
        data.pop("config_source", None)
        # Add helpful comments as leading keys
        commented = {
            "_comment_1": "Sourcing filter configuration — edit values below to control approvals.",
            "_comment_2": "Set a filter to null or 0 to disable it. Percentages are fractions (0.15 = 15%).",
            **data,
        }
        Path(path).write_text(
            json.dumps(commented, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


@dataclass
class SourcingRow:
    # --- Zoro/supplier data ---
    status: str = ""
    zoro_sku: str | None = None
    zoro_title: str | None = None
    brand: str | None = None
    model: str | None = None
    upc: str | None = None
    zoro_cost: float | None = None            # ALWAYS in USD after conversion
    supplier_original_price: float | None = None  # raw scraped price
    supplier_currency: str = "USD"            # what currency the original was in
    supplier_to_usd_rate: float = 1.0
    # --- Amazon data ---
    amazon_asin: str | None = None
    amazon_sell_price: float | None = None
    fee_pct: float = 0.15
    fee_dollars: float | None = None
    extra_cost: float = 0.0
    profit_dollars: float | None = None
    margin_pct: float | None = None
    roi_pct: float | None = None
    # Seller counts — all of them, client can filter later
    fbm_sellers: Any = NEEDS_KEEPA           # live FBM count (client's primary rule)
    fba_sellers: Any = NEEDS_KEEPA           # live FBA count
    total_sellers_live: Any = NEEDS_KEEPA    # live FBA + FBM
    historical_sellers: Any = NEEDS_KEEPA    # every seller ever seen — market churn signal
    amazon_on_listing: str = "?"             # "Yes" / "No" / "?" (unknown)
    rating: str | None = None
    reviews: str | None = None
    bsr: Any = NEEDS_KEEPA
    bsr_avg_90d: Any = NEEDS_KEEPA           # 90-day avg BSR (velocity signal)
    category: str | None = None
    buy_box_price: float | None = None
    ships_from: str | None = None            # Amazon seller name winning the Buy Box
    avg_price_90d: float | None = None
    # --- Decision ---
    reject_reasons: str = ""
    # --- Extras (not in demo but useful) ---
    zoro_url: str | None = None
    amazon_url: str | None = None
    search_query: str | None = None
    amazon_title: str | None = None
    availability: str | None = None  # supplier stock status (InStock / OutOfStock / …)


# ---------- Building rows ------------------------------------------------- #

def build_sourcing_rows_from_supplier(
    supplier_products: list[dict],
    cfg: SourcingConfig,
) -> list[SourcingRow]:
    """
    Build sourcing rows straight from supplier scrape data — no Amazon lookup yet.
    Used with Keepa-based enrichment where Amazon data comes from the API,
    not the fragile Amazon web scraper.

    UPC is preserved for the Keepa lookup that follows.
    """
    rows: list[SourcingRow] = []
    rate = cfg.supplier_to_usd_rate if cfg.supplier_to_usd_rate > 0 else 1.0

    for p in supplier_products:
        specs = p.get("specs") or {}
        upc = (specs.get("UPC") or specs.get("gtin13") or p.get("upc")
               or "").strip() or None
        model = specs.get("Model") or specs.get("MFR #") or specs.get("mpn")

        original_price = _to_float(p.get("price"))
        # Convert to USD if a non-1 rate was configured
        usd_cost = None
        if original_price is not None:
            usd_cost = round(original_price / rate, 2) if rate != 1.0 else original_price

        row = SourcingRow(
            zoro_sku=p.get("sku"),
            zoro_title=p.get("title"),
            brand=p.get("brand"),
            model=model,
            upc=upc,
            zoro_cost=usd_cost,
            supplier_original_price=original_price,
            supplier_currency=cfg.supplier_currency,
            supplier_to_usd_rate=rate,
            zoro_url=p.get("url"),
            availability=p.get("availability"),
            fee_pct=cfg.fee_pct,
            extra_cost=cfg.extra_cost,
        )
        rows.append(row)
    return rows


def _apply_keepa_match(row: SourcingRow, match) -> None:
    """Populate Amazon-side fields on `row` from a KeepaMatch."""
    row.amazon_asin = match.asin
    row.amazon_title = match.title
    row.amazon_url = f"https://www.amazon.com/dp/{match.asin}"
    row.amazon_sell_price = match.buy_box_price or match.new_price
    row.buy_box_price = match.buy_box_price
    row.avg_price_90d = match.avg_price_90d
    row.fbm_sellers = match.fbm_seller_count_live
    row.fba_sellers = match.fba_seller_count_live
    row.total_sellers_live = match.total_offer_count_live
    row.historical_sellers = match.total_historical_sellers
    row.amazon_on_listing = "Yes" if match.amazon_on_listing else "No"
    row.rating = str(match.rating) if match.rating else None
    row.reviews = str(match.review_count) if match.review_count else None
    row.bsr = match.bsr
    row.bsr_avg_90d = match.bsr_avg_90d
    row.category = match.category
    # Buy Box seller name is populated inline when Amazon has the box; other
    # cases are deferred to a batched /seller lookup by the caller.
    if match.buy_box_price is not None and match.buy_box_seller_name:
        row.ships_from = match.buy_box_seller_name


def _clone_for_variation(source: SourcingRow) -> SourcingRow:
    """
    Create a fresh SourcingRow carrying the supplier-side data from `source`
    but with Amazon-side fields reset to defaults, so a second/third ASIN
    variation gets its own row rather than overwriting the first.
    """
    fresh = SourcingRow()
    # Copy every supplier / config field; leave Amazon fields at defaults.
    supplier_attrs = (
        "zoro_sku", "zoro_title", "brand", "model", "upc",
        "zoro_cost", "supplier_original_price",
        "supplier_currency", "supplier_to_usd_rate",
        "fee_pct", "extra_cost",
        "zoro_url", "availability", "search_query",
    )
    for attr in supplier_attrs:
        setattr(fresh, attr, getattr(source, attr))
    return fresh


def enrich_rows_with_keepa(
    rows: list[SourcingRow],
    keepa_client,
    cfg: "SourcingConfig | None" = None,
) -> list[SourcingRow]:
    """
    Take sourcing rows produced from supplier scraping, look up each UPC on
    Keepa, and fill in the Amazon-side fields (ASIN, BSR, seller counts, etc.).

    IMPORTANT: A single supplier product may map to multiple Amazon ASINs
    (variations: sizes, colours, pack counts — all sharing one UPC family).
    This function fans out — each qualifying ASIN gets its own row in the
    returned list, sharing the same supplier data but with independent
    Amazon-side data and independent rule evaluation. That means the
    output list can be LONGER than the input list.

    Rows without a UPC are marked "no UPC" and skipped.
    Rows where Keepa finds no match are marked "no Amazon match".
    """
    # Track (row, seller_id) pairs so we can batch-resolve names after the
    # product loop instead of paying one HTTP round-trip per product.
    pending_seller_lookup: list[tuple[SourcingRow, str]] = []
    enriched: list[SourcingRow] = []
    total_variations = 0
    total_via_search_only = 0

    # Title search fallback catches Amazon listings with no UPC (very common
    # for pack-size variants). Adds ~10 tokens per row but roughly doubles
    # ASIN coverage. Can be disabled by setting KEEPA_TITLE_SEARCH=0 in env.
    import os as _os
    title_search_enabled = (
        _os.environ.get("KEEPA_TITLE_SEARCH", "1").strip().lower()
        not in ("0", "false", "no", "off")
        and hasattr(keepa_client, "search_by_brand_title")
    )

    # Import lazily so keepa.py stays importable in unit-test contexts where
    # `requests` isn't installed.
    try:
        from keepa import KeepaTokenExhausted
    except ImportError:  # pragma: no cover
        class KeepaTokenExhausted(Exception):  # type: ignore
            refill_seconds = None

    tokens_exhausted = False  # sticky flag: once tripped, skip remaining Keepa calls
    exhausted_refill_s: int | None = None

    for i, row in enumerate(rows, 1):
        # If tokens ran out on an earlier row, don't hit Keepa again —
        # mark this row INCOMPLETE and keep going. The job still finishes
        # with a valid Excel; the rows we DID complete are all real data.
        if tokens_exhausted:
            reason = "Keepa tokens exhausted mid-run"
            if exhausted_refill_s:
                mins = max(1, exhausted_refill_s // 60)
                reason += f" — refills in ~{mins}min"
            row.reject_reasons = _append_reason(row.reject_reasons, reason)
            row.status = "INCOMPLETE"
            enriched.append(row)
            continue

        # STEP 1 — UPC lookup (fast, exact when it works).
        upc_matches = []
        if row.upc:
            log.info("[%d/%d] Keepa UPC %s", i, len(rows), row.upc)
            try:
                upc_matches = keepa_client.lookup_all_by_upc(row.upc)
            except KeepaTokenExhausted as e:
                log.error("Keepa tokens exhausted at row %d/%d — saving "
                          "partial results and marking remaining rows "
                          "INCOMPLETE", i, len(rows))
                tokens_exhausted = True
                exhausted_refill_s = getattr(e, "refill_seconds", None)
                reason = f"Keepa tokens exhausted before this row: {e}"
                row.reject_reasons = _append_reason(row.reject_reasons, reason)
                row.status = "INCOMPLETE"
                enriched.append(row)
                continue
            except Exception as e:  # noqa: BLE001
                log.warning("Keepa UPC error for %s: %s", row.upc, e)
                row.reject_reasons = _append_reason(
                    row.reject_reasons, f"Keepa UPC error: {e}"
                )
                row.status = "INCOMPLETE"
                enriched.append(row)
                continue

        # STEP 2 — title/brand search fallback. Catches:
        #   * Products where iHerb has no UPC in JSON-LD
        #   * Amazon ASINs with UPC list = [] (seller never registered a
        #     barcode) — e.g. B0BLHRM711 Alter Eco Granola 6-pack
        #   * Additional variations Amazon linked under different UPCs
        search_matches = []
        if title_search_enabled and row.brand and row.zoro_title:
            if not row.upc:
                log.info("[%d/%d] Keepa title search '%s' (no UPC)",
                         i, len(rows), row.brand)
            try:
                search_matches = keepa_client.search_by_brand_title(
                    brand=row.brand, title=row.zoro_title, limit=8,
                )
            except KeepaTokenExhausted as e:
                log.warning("Keepa tokens exhausted during title search — "
                            "continuing with UPC results only for remainder")
                tokens_exhausted = True
                exhausted_refill_s = getattr(e, "refill_seconds", None)
                # Do NOT continue — we still have UPC matches to process
            except Exception as e:  # noqa: BLE001
                log.warning("Keepa title search error for '%s' / '%s': %s",
                            row.brand, (row.zoro_title or "")[:40], e)

        # Merge UPC + search results, dedupe by ASIN (UPC hits first so
        # they win rank order — they're the strongest identity signal).
        seen_asins: set[str] = set()
        matches = []
        for m in upc_matches + search_matches:
            if m.asin and m.asin not in seen_asins:
                seen_asins.add(m.asin)
                matches.append(m)

        # Defensive cap so a bad brand+title combo can't fan out infinitely.
        matches = matches[:15]

        if not matches:
            if not row.upc and not row.brand:
                reason = "no UPC and no brand/title on supplier product"
            elif row.upc:
                reason = "no Amazon match (tried UPC + title search)"
            else:
                reason = "no Amazon match for brand/title"
            row.reject_reasons = _append_reason(row.reject_reasons, reason)
            row.status = "Rejected"
            enriched.append(row)
            continue

        # Diagnostic: how many ASINs came from search that UPC alone missed?
        upc_asins = {m.asin for m in upc_matches}
        search_only = [m for m in search_matches if m.asin not in upc_asins]
        if search_only and not upc_matches:
            total_via_search_only += len(search_only)
            log.info("  → title search rescued %d ASIN(s) (no UPC match): %s",
                     len(search_only), ", ".join(m.asin for m in search_only[:3]))
        elif search_only:
            total_via_search_only += len(search_only)
            log.info("  → title search added %d extra ASIN(s) on top of UPC: %s",
                     len(search_only), ", ".join(m.asin for m in search_only[:3]))

        if len(matches) > 1:
            total_variations += len(matches) - 1
            log.info("  → %d Amazon variations found; fanning out to "
                     "one row per ASIN", len(matches))

        for idx, match in enumerate(matches):
            # First match reuses the incoming row (preserves any pre-set
            # supplier fields untouched); additional variations get a clone
            # so each ASIN has independent Amazon data and its own status.
            target = row if idx == 0 else _clone_for_variation(row)

            _apply_keepa_match(target, match)

            if (match.buy_box_price is not None
                    and not target.ships_from
                    and match.buy_box_seller_id):
                pending_seller_lookup.append((target, match.buy_box_seller_id))

            _compute_profitability(target)
            if cfg is not None:
                _apply_rules(target, cfg)

            # Preserves rank order: best-scored match (idx=0) appears first,
            # subsequent variations in ranked order below it.
            enriched.append(target)

        # Live token-balance heartbeat every 10 rows or so. Gives operators
        # a real-time signal of how fast the quota is draining without
        # spamming the log.
        tokens = getattr(keepa_client, "tokens_left", None)
        if tokens is not None and (i % 10 == 0 or tokens < 100):
            log.info("[%d/%d] Keepa tokens remaining: %d", i, len(rows), tokens)

    if total_variations:
        log.info("Fanned out %d extra Amazon variation row(s) across %d supplier product(s)",
                 total_variations, len(rows))
    if total_via_search_only:
        log.info("Title-search fallback surfaced %d ASIN(s) that UPC lookup would have missed",
                 total_via_search_only)

    # Batch-resolve all Buy Box seller names in one /seller call (up to 100
    # unique IDs per call — very few unique sellers usually, so this is cheap).
    if pending_seller_lookup and hasattr(keepa_client, "lookup_seller_names"):
        unique_ids = list({sid for _, sid in pending_seller_lookup})
        try:
            names = keepa_client.lookup_seller_names(unique_ids)
        except Exception as e:  # noqa: BLE001
            log.warning("Keepa seller-name batch lookup failed: %s", e)
            names = {}
        for r, sid in pending_seller_lookup:
            # Prefer real name; fall back to raw sellerId so the column is
            # never mysteriously blank when we know a Buy Box seller exists.
            r.ships_from = names.get(sid) or sid

    # Final status summary — surfaces prominently in job logs and the UI.
    tokens = getattr(keepa_client, "tokens_left", None)
    if tokens is not None:
        log.info("Keepa tokens remaining: %d", tokens)
    if tokens_exhausted:
        incomplete_count = sum(1 for r in enriched if r.status == "INCOMPLETE"
                               and "Keepa tokens exhausted" in (r.reject_reasons or ""))
        completed_count = len(enriched) - incomplete_count
        refill_msg = ""
        if exhausted_refill_s:
            mins = max(1, exhausted_refill_s // 60)
            refill_msg = f" (Keepa refills in ~{mins}min — retry then)"
        log.warning("KEEPA QUOTA EXHAUSTED: completed %d rows fully, %d rows "
                    "marked INCOMPLETE.%s", completed_count, incomplete_count,
                    refill_msg)

    return enriched


def _append_reason(existing: str, new: str) -> str:
    if not existing:
        return new
    return f"{existing}; {new}"


def _to_float(x: Any) -> float | None:
    if x is None or x == "":
        return None
    if isinstance(x, (int, float)):
        return float(x)
    import re
    m = re.search(r"[\d,]+\.?\d*", str(x))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def build_sourcing_rows(
    zoro_products: list[dict],
    searcher: AmazonSearcher,
    cfg: SourcingConfig,
    enrich_top_match: bool = True,
) -> list[SourcingRow]:
    """
    For each Zoro product:
      1. Amazon search → pick top match
      2. Optionally fetch product page for ASIN → get category, BSR-if-visible,
         sold-by (for Amazon-on-listing), UPC, model
      3. Compute profitability
      4. Apply sourcing rules → status + reject reasons
    """
    # Step 1: search
    comparison = compare_zoro_to_amazon(zoro_products, searcher)

    rows: list[SourcingRow] = []
    for i, (zp, comp) in enumerate(zip(zoro_products, comparison), 1):
        row = SourcingRow(
            zoro_sku=zp.get("sku"),
            zoro_title=zp.get("title"),
            brand=zp.get("brand"),
            zoro_cost=_to_float(zp.get("price")),
            zoro_url=zp.get("url"),
            search_query=comp.search_query,
            availability=zp.get("availability"),
            fee_pct=cfg.fee_pct,
            extra_cost=cfg.extra_cost,
        )

        # Pick the top Amazon match (first non-sponsored, else first)
        top: AmazonResult | None = None
        if comp.amazon_results:
            non_sponsored = [r for r in comp.amazon_results if not r.sponsored]
            top = non_sponsored[0] if non_sponsored else comp.amazon_results[0]

        if top:
            row.amazon_asin = top.asin
            row.amazon_sell_price = _to_float(top.price)
            row.rating = top.rating
            row.reviews = top.review_count
            row.amazon_url = top.url
            row.amazon_title = top.title

            # Step 2: enrich from product page (optional)
            if enrich_top_match and top.asin:
                log.info("[%d/%d] Enriching ASIN %s", i, len(zoro_products), top.asin)
                details = searcher.fetch_product_details(top.asin)
                _merge_details(row, details)

        # Step 3: profitability
        _compute_profitability(row)

        # Step 4: apply rules
        _apply_rules(row, cfg)

        rows.append(row)
    return rows


def _merge_details(row: SourcingRow, d: AmazonProductDetails) -> None:
    """Merge product-page details into the sourcing row."""
    if d.captcha_hit:
        return  # keep whatever we had from search
    if d.title:
        row.amazon_title = d.title
    if d.price:
        row.amazon_sell_price = _to_float(d.price)
    if d.rating:
        row.rating = d.rating
    if d.review_count:
        row.reviews = d.review_count
    if d.category:
        row.category = d.category
    if d.upc and not row.upc:
        row.upc = d.upc
    if d.model and not row.model:
        row.model = d.model
    if d.bsr:
        row.bsr = d.bsr  # override the KEEPA placeholder if we managed to scrape it
    if d.amazon_on_listing is not None:
        row.amazon_on_listing = "Yes" if d.amazon_on_listing else "No"


def _compute_profitability(row: SourcingRow) -> None:
    if row.amazon_sell_price is None or row.zoro_cost is None:
        return
    sell = row.amazon_sell_price
    cost = row.zoro_cost
    fee = round(sell * row.fee_pct, 2)
    profit = round(sell - fee - row.extra_cost - cost, 2)

    row.fee_dollars = fee
    row.profit_dollars = profit
    row.margin_pct = round(profit / sell, 6) if sell > 0 else None
    row.roi_pct = round(profit / cost, 6) if cost > 0 else None


def _apply_rules(row: SourcingRow, cfg: SourcingConfig) -> None:
    reasons: list[str] = []

    # === Supplier stock ===
    if cfg.require_in_stock and is_out_of_stock(row.availability):
        label = row.availability or "OutOfStock"
        reasons.append(f"supplier out of stock ({label})")

    # === Competition rules ===

    # Rule: Amazon on the listing (client's rule #1)
    if cfg.reject_if_amazon_on_listing and row.amazon_on_listing == "Yes":
        reasons.append("Amazon is on the listing")

    # Rule: FBM sellers — min (client's ≥4 rule)
    if isinstance(row.fbm_sellers, (int, float)):
        if row.fbm_sellers < cfg.min_fbm_sellers:
            reasons.append(f"only {int(row.fbm_sellers)} FBM sellers (need ≥{cfg.min_fbm_sellers})")
        elif cfg.max_fbm_sellers is not None and row.fbm_sellers > cfg.max_fbm_sellers:
            reasons.append(f"{int(row.fbm_sellers)} FBM sellers (max {cfg.max_fbm_sellers} — too crowded)")

    # Rule: historical sellers (validates "mature listing" if set)
    if cfg.min_historical_sellers > 0 and isinstance(row.historical_sellers, (int, float)):
        if row.historical_sellers < cfg.min_historical_sellers:
            reasons.append(
                f"only {int(row.historical_sellers)} historical sellers "
                f"(need ≥{cfg.min_historical_sellers} for market validation)"
            )

    # === Product quality rules ===

    # Rule: Rating floor (client's ≥3.5 rule)
    rating_value = _to_float(row.rating)
    if rating_value is not None:
        if rating_value < cfg.min_rating:
            reasons.append(f"rating {rating_value} below {cfg.min_rating}")
    elif cfg.require_rating:
        reasons.append("no rating data (require_rating is on)")

    # Rule: Review count
    review_value = _to_float(row.reviews)
    if review_value is not None:
        if cfg.min_reviews > 0 and review_value < cfg.min_reviews:
            reasons.append(f"only {int(review_value)} reviews (need ≥{cfg.min_reviews})")
    elif cfg.require_reviews:
        reasons.append("no reviews (require_reviews is on)")

    # === Velocity rules ===

    # Rule: BSR ceiling (lower BSR = better; high BSR = slow mover)
    if cfg.max_bsr is not None and isinstance(row.bsr, (int, float)):
        if row.bsr > cfg.max_bsr:
            reasons.append(f"BSR {int(row.bsr):,} above max {cfg.max_bsr:,} (slow mover)")

    # === Profitability rules ===

    if row.profit_dollars is not None:
        if row.profit_dollars < cfg.min_profit:
            reasons.append(f"profit ${row.profit_dollars:.2f} below ${cfg.min_profit:.2f}")
        elif row.margin_pct is not None and row.margin_pct < cfg.min_margin:
            reasons.append(f"margin {row.margin_pct:.1%} below {cfg.min_margin:.1%}")
        elif row.roi_pct is not None and row.roi_pct < cfg.min_roi:
            reasons.append(f"ROI {row.roi_pct:.1%} below {cfg.min_roi:.1%}")

    # === Match rules ===

    # If we have no Amazon match at all
    if row.amazon_asin is None:
        reasons.append("no Amazon match found")

    row.reject_reasons = "; ".join(reasons)

    if row.reject_reasons:
        row.status = "Rejected"
    elif row.amazon_asin and row.profit_dollars is not None:
        # We have data and no reasons to reject
        # If Keepa fields are missing, mark REVIEW so user knows to verify
        keepa_missing = row.fbm_sellers == NEEDS_KEEPA or row.bsr == NEEDS_KEEPA
        row.status = "REVIEW" if keepa_missing else "APPROVED"
    else:
        row.status = "INCOMPLETE"


# ---------- Output -------------------------------------------------------- #

DEMO_COLUMNS = [
    ("status",                  "Status"),
    ("zoro_sku",                "Supplier SKU"),
    ("zoro_title",              "Supplier Title"),
    ("availability",            "Supplier Stock"),
    ("brand",                   "Brand"),
    ("model",                   "Model"),
    ("upc",                     "UPC"),
    ("supplier_original_price", "Supplier Cost (Original)"),
    ("supplier_currency",       "Original Currency"),
    ("supplier_to_usd_rate",    "FX Rate (per USD)"),
    ("zoro_cost",               "Supplier Cost (USD)"),
    ("amazon_asin",             "Amazon ASIN"),
    ("amazon_title",            "Amazon Title"),
    ("amazon_sell_price",       "Amazon Sell Price"),
    ("buy_box_price",           "Buy Box Price"),
    ("ships_from",              "Ships From"),
    ("avg_price_90d",           "Avg Price 90d"),
    ("fee_pct",                 "Fee %"),
    ("fee_dollars",             "Fee $"),
    ("extra_cost",              "Ship $ / unit"),
    ("profit_dollars",          "Profit $"),
    ("margin_pct",              "Margin %"),
    ("roi_pct",                 "ROI %"),
    ("fbm_sellers",             "FBM Sellers (live)"),
    ("fba_sellers",             "FBA Sellers (live)"),
    ("total_sellers_live",      "Total Sellers (live)"),
    ("historical_sellers",      "Historical Sellers"),
    ("amazon_on_listing",       "Amazon on Listing"),
    ("rating",                  "Rating"),
    ("reviews",                 "Reviews"),
    ("bsr",                     "BSR"),
    ("bsr_avg_90d",             "BSR Avg 90d"),
    ("category",                "Category"),
    ("reject_reasons",          "Reject Reasons"),
]


def _row_to_dict(row: SourcingRow) -> dict:
    return {label: getattr(row, attr) for attr, label in DEMO_COLUMNS}


def save_sourcing_csv(rows: list[SourcingRow], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[label for _, label in DEMO_COLUMNS])
        w.writeheader()
        for r in rows:
            w.writerow(_row_to_dict(r))
    log.info("Wrote %d sourcing rows to %s", len(rows), path)


def save_sourcing_xlsx(rows: list[SourcingRow], path: Path) -> None:
    """Write an .xlsx matching the demo template layout with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        log.error("Falling back to CSV: %s", path.with_suffix(".csv"))
        save_sourcing_csv(rows, path.with_suffix(".csv"))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "sourcing"

    # Row 1: descriptive title (matches demo)
    ws["A1"] = ("Zoro -> Amazon Sourcing  |  Zoro data = REAL  |  "
                "NEEDS_KEEPA cells require Keepa/paid API to fill")
    ws["A1"].font = Font(bold=True, italic=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DEMO_COLUMNS))

    # Row 2: headers
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (_, label) in enumerate(DEMO_COLUMNS, 1):
        c = ws.cell(row=2, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Colour fills for status
    fills = {
        "APPROVED":   PatternFill("solid", fgColor="C6EFCE"),
        "REVIEW":     PatternFill("solid", fgColor="FFEB9C"),
        "Rejected":   PatternFill("solid", fgColor="FFC7CE"),
        "INCOMPLETE": PatternFill("solid", fgColor="D9D9D9"),
    }

    # Data rows
    for r_idx, row in enumerate(rows, 3):
        data = _row_to_dict(row)
        for c_idx, (attr, label) in enumerate(DEMO_COLUMNS, 1):
            val = data[label]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Format numbers
            if attr in ("zoro_cost", "amazon_sell_price", "buy_box_price",
                        "avg_price_90d", "fee_dollars",
                        "extra_cost", "profit_dollars"):
                cell.number_format = '"$"#,##0.00'
            elif attr == "supplier_original_price":
                cell.number_format = '#,##0.00'
            elif attr == "supplier_to_usd_rate":
                cell.number_format = '#,##0.0000'
            elif attr in ("margin_pct", "roi_pct", "fee_pct"):
                cell.number_format = "0.00%"
        # Status colouring
        status_cell = ws.cell(row=r_idx, column=1)
        if status_cell.value in fills:
            status_cell.fill = fills[status_cell.value]
            status_cell.font = Font(bold=True)

    # Column widths — approximate the demo
    widths = {
        "Status": 12, "Supplier SKU": 14, "Supplier Title": 55,
        "Supplier Stock": 16, "Brand": 16,
        "Model": 14, "UPC": 14,
        "Supplier Cost (Original)": 14, "Original Currency": 10, "FX Rate (per USD)": 12,
        "Supplier Cost (USD)": 14,
        "Amazon ASIN": 13,
        "Amazon Title": 40, "Amazon Sell Price": 13, "Buy Box Price": 12,
        "Ships From": 22,
        "Avg Price 90d": 12, "Fee %": 8, "Fee $": 10,
        "Ship $ / unit": 12, "Profit $": 10, "Margin %": 10, "ROI %": 10,
        "FBM Sellers (live)": 14, "FBA Sellers (live)": 14,
        "Total Sellers (live)": 16, "Historical Sellers": 15,
        "Amazon on Listing": 15, "Rating": 8, "Reviews": 10,
        "BSR": 12, "BSR Avg 90d": 12,
        "Category": 22, "Reject Reasons": 40,
    }
    from openpyxl.utils import get_column_letter
    for c_idx, (_, label) in enumerate(DEMO_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(label, 14)

    ws.freeze_panes = "A3"
    wb.save(path)
    log.info("Wrote %d sourcing rows to %s", len(rows), path)
